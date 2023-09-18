from ast import keyword
from collections import Counter
from itertools import groupby
from pathlib import Path
from unicodedata import name
from openpyxl import load_workbook
from diff_folder_and_mysql import get_sheet_structure 
import pyodbc 
import os
import shutil
from pymongo_get_database import get_database
       



def all_equal(iterable):
    g = groupby(iterable)
    return next(g, True) and not next(g, False)

def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()

def indices(lst, item):
    return [i for i, x in enumerate(lst) if x == item]

def get_entry_by_name(sheet, table_name, names, columns,ranges, row):
    entries = []
    table_indices = indices(names, table_name)
    for index in table_indices:
        entry = []
        styles = []
        min_col, min_row,max_col,max_row=ranges[index]
        range = (min_col, row, max_col, row) 
        for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
            for cell in cells:
                entry.append(cell.value)
                styles.append(type(cell.value))
        entries.append((entry,range,columns[index],styles))
    return entries


def work_with_line_in_structure_lessons(line,course_sheet, course_path, cursor):
    path=course_path
    sheet = course_sheet
    folder_col = get_column(sheet=sheet, row=1, name='Folders')
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    min_col, min_row,max_col,max_row=ranges[names.index('WrapperExercises')]
    ids = columns[names.index('WrapperExercises')].index('Exercise_Id')
    
    entry = []
    for cells in sheet.iter_cols(min_col=min_col,min_row=line, max_col=max_col, max_row=line):
        for cell in cells:
            entry.append(cell.value)

    id = entry[columns[names.index('WrapperExercises')].index('Exercise_Id')]
    min_col, min_row,max_col,max_row=ranges[names.index('Level names')]
    level = []
    for cells in sheet.iter_cols(min_col=min_col,min_row=line, max_col=max_col, max_row=line):
        for cell in cells:
            level.append(cell.value)
    print(level)
    dick= {}
    dick['MinimalPair_UPSIDSound']=""
    dick['Second_MinimalPair_UPSIDSound']=""
    dick['MinimalPair_IPA']=""
    dick['Second_MinimalPair_IPA']=""

#    if id !=None:
#        dick['_id'] = str(id)
    if level[-1] != None:
        print(level[-1]) 
        dick["Exercise_name"] =  level[-1]
        #find group name in this case
        print(min_col, min_row,max_col,max_row)
        value = False
        current_row = line-1
        while not value:
            cell = sheet.cell(current_row,column=max_col-1)
            if cell.value:
                value = True
                print(cell.value, 'Group_exercise')
                dick['Group_lesson'] = cell.value
            current_row -= 1
    elif level[-2] != None:
        #dick["Exercise name"] =  level[-2]
        pass
    else:
        pass

   
    value = False
    current_row = line-1
    while not value:
        cell = sheet.cell(current_row,column=max_col-2)
        if cell.value:
            value = True
            print(cell.value, 'lesson')
            dick['Lesson'] = cell.value
        current_row -= 1

    value = False
    current_row = line-1
    while not value:
        cell = sheet.cell(current_row,column=max_col-3)
        if cell.value:
            value = True
            print(cell.value, 'level')
            dick['Level'] = cell.value
        current_row -= 1
    
    #open an exercise.xlsx
    exercise_path = Path(sheet.cell(row=line, column=folder_col).value.replace('..',str(path.parent)))
    exercise_file = exercise_path/'exercise.xlsx'
    print(exercise_file.exists(), str(exercise_file))
    
    wb_exercise = load_workbook(str(exercise_file))

    data_row = 3
    sheet_ex = wb_exercise['Exercise']
    names,columns,ranges = get_sheet_structure(sheet = sheet_ex)
    c_prop = Counter(names)
    print(c_prop)
    
    for i in range(c_prop['Properties']):
            #it creates an entry with proper exercise id and id if it doesnt exist in db, excel file gets upodated, pay attention, 
            # exercises.xlsx shoul be closed
            
        entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_ex,table_name='Properties', names=names,ranges=ranges,row=data_row,columns=columns)[i]
        print(entry, entry[entry_columns.index('Key')])
        if entry[entry_columns.index('Key')]:
            if entry[entry_columns.index('Value')]:
                dick[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
            if entry[entry_columns.index('Description')]:
                dick[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
    if dick['MinimalPair_UPSIDSound'] !="":
        va = str(dick['MinimalPair_UPSIDSound'])
        
        if '/' in va:
            for lue in va.split('/'):
                print(lue)
                sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Sounds] where Id = '+lue
                cursor.execute(sqlcommand)
                sound = cursor.fetchall()[0]
                dick['MinimalPair_UPSIDSound'] = dick['MinimalPair_UPSIDSound'].replace(lue, sound[4])
                dick['MinimalPair_IPA'] = dick['MinimalPair_UPSIDSound'].replace(sound[4],sound[1])

        else:
            sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Sounds] where Id = '+str(dick['MinimalPair_UPSIDSound'])
            cursor.execute(sqlcommand)
            sound = cursor.fetchall()[0]
            dick['MinimalPair_UPSIDSound'] = sound[4]
            dick['MinimalPair_IPA'] = sound[1]
    else:
        pass
    if dick['Second_MinimalPair_UPSIDSound'] != "":
        va = str(dick['Second_MinimalPair_UPSIDSound'])
        if '/' in va:
            for lue in va.split('/'):
                print(lue)
                sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Sounds] where Id = '+lue
                cursor.execute(sqlcommand)
                sound = cursor.fetchall()[0]
                dick['Second_MinimalPair_UPSIDSound'] = dick['Second_MinimalPair_UPSIDSound'].replace(lue, sound[4])
                dick['Second_MinimalPair_IPA'] = dick['Second_MinimalPair_UPSIDSound'].replace(sound[4],sound[1])

        else:
            sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Sounds] where Id = '+str(dick['Second_MinimalPair_UPSIDSound'])
            cursor.execute(sqlcommand)
            sound = cursor.fetchall()[0]
            dick['Second_MinimalPair_UPSIDSound'] = sound[4]
            dick['Second_MinimalPair_IPA'] = sound[1]
    else:
        pass

        
        

    
    case_vocab = False
    vocab = []
    case_mp = False
    mp = []
    case_nw = False
    nw = []
    print(wb_exercise.sheetnames, "mitya")
    
    
    for sheetname in wb_exercise.sheetnames:
        print(sheetname)
        if 'Vocab' in  sheetname:
            case_vocab = False
            vocab.append(sheetname)
        elif 'MP' in sheetname:
            #work on Vocab-Confusion Box
            case_mp = True
            mp.append(sheetname)
        elif 'NonWords' in sheetname:
            case_nw = True
            nw.append(sheetname)
        else:
            pass
  
    print(case_mp,case_vocab, case_nw)
    dick['MP_wordpairs']=[]
    dick['MP_nonwords'] = []
    
    

    if case_vocab:
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = False
        print(vocab)
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in vocab:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        for row in range(data_row,maximum_row+1):
            dick_word = {}
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
          
             #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            print(entry)
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            print(dick_word)
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
                
            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
     
            words.append(dick_word)

     
        dick['Vocabulary_words'] = words
        
        #with open('result.json', 'w') as fp:
        #    json.dump(dick, fp)
        
    if case_mp:

        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        
        vocab = mp
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in mp:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        switch = 0
        wordpair = []
        for row in range(data_row,maximum_row+1):
            if switch == 0:
                wordpair = []
            else:
                pass

            dick_word = {}
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
        
            #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            print(entry)
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            print(dick_word)
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
                
            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
            wordpair.append(dick_word)
            if switch == 0:
                switch = 1
            else:
                switch = 0
                words.append(wordpair)

        
        
        dick['MP_wordpairs'] = words
                
   
        
     
    if case_nw:
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = False
        vocab = nw
        print(vocab)
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in vocab:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        
        sheet_cb = wb_exercise[sheet_name]
        names_cb,columns_cb,ranges_cb = get_sheet_structure(sheet = sheet_cb)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        for row in range(data_row,maximum_row+1):
            dick_word = {}
            #start with confusionbox,

            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
          
             #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            print(entry)
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            print(dick_word)
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]



            c = Counter(names_cb)
            print(c['Properties'], "here", names_cb)
            
            if c['Properties']==1:

                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_cb,table_name='Properties', names=names_cb,ranges=ranges_cb,row=row,columns=columns_cb)[-1]

                print(entry)
                if entry[entry_columns.index('Key')] == 'PairWord':
                    dick_word['nonword'] = entry[entry_columns.index('Value')]
                else:
                    pass
                
                
            
            c = Counter(names_wp)
            print(c['Properties'], "here")



            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
     
            words.append(dick_word)

     
        dick['MP_nonwords'] = words
        
        #import json
        #with open('result.json', 'w') as fp:
        #    json.dump(dick, fp)
         
   

    else:
        pass

    return(dick)
 
               
def main(cursor, collection_name, course_folder):
    path = Path(course_folder)
    #the path to the course summary file
    structure_file  = Path(course_folder+'\\lessons_structure.xlsx')

    if structure_file.exists():
        #check actions column for the command "submit"
        wb_structure = load_workbook(str(structure_file))
        sheet = wb_structure['Lessons']
        to_submit = []
        print(str(structure_file))
        action_col = get_column(sheet=sheet, row = 2, name='Level 4')
    
    
        key = get_column(sheet=sheet, row = 1, name='Actions')
        print(key, action_col)
        
        for cells in sheet.iter_cols(min_col=action_col,min_row=3, max_col=action_col, max_row=sheet.max_row):
            for cell in cells:
                #if str(cell.value).lower() == 'test':
                if cell.value and not sheet.cell(row=cell.row, column=key).value == 'del':
                    
                    to_submit.append(cell.row)
                    print(cell.value)
        # list to_submit contains rows of summary file to submit
        print('rows to_submit: ', to_submit)
        
        
        
        
        for line in to_submit:
            cell_action = sheet.cell(row=line,column=action_col)
            #try:
            dictionary = work_with_line_in_structure_lessons(cursor=cursor, line=line,course_sheet=sheet,course_path=path)
       
            
            collection_name.insert_one(dictionary)
            
            #cell_action.value = 'submitted'
            #wb_structure.save(structure_file)
            #except:
                #cell_action.value = 'failed_submitted'
                #wb_structure.save(structure_file)  
                     
        
    else:
        print("No exercise file here. quitting")
        exit() 

if __name__ == "__main__":
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    
    #connect to the calst database
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    language='Norwegian'
    dbname = get_database('calst_test')
    collection_name = dbname[language]
    #collection_name = dbname['italian_test_nonword']
    output_sound = r'C:\Source\Repos\CalstEnglish\CalstFiles\WordObjectContent'+'\\'+language+r'\OriginalWords_Wav'
    print(output_sound)
 
    keyword = 'retract'	
    dst_path = Path(output_sound)
    #folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\'
    #folder = 'G:\\My Drive\\CALST_courses\\'+str(language)+'_course_styled\\'
    #folder = 'C:\\Source\\Repos\\mysql-excel\\'+str(language)+'_course_styled\\'
    #folder = r'C:\Users\dmitrysh\OneDrive - NTNU\CALST_courses\\'+str(language)+'_course_styled'
    #folder = r'C:\Users\dmitrysh\OneDrive - NTNU'
    folder = r'C:\Source\Repos\mysql-excel\Norwegian_course_revised'
    main(cursor=cursor, collection_name = collection_name, course_folder=folder)
    #cnxn.commit()
    #cnxn.close()