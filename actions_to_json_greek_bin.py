from ast import keyword
from collections import Counter
from itertools import groupby
from pathlib import Path
from unicodedata import name
from openpyxl import load_workbook
from diff_folder_and_mysql import get_sheet_structure 
import pyodbc 
import os
import shutil
from pymongo_get_database import get_database
       



def all_equal(iterable):
    g = groupby(iterable)
    return next(g, True) and not next(g, False)

def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()

def indices(lst, item):
    return [i for i, x in enumerate(lst) if x == item]

def get_entry_by_name(sheet, table_name, names, columns,ranges, row):
    entries = []
    table_indices = indices(names, table_name)
    for index in table_indices:
        entry = []
        styles = []
        min_col, min_row,max_col,max_row=ranges[index]
        range = (min_col, row, max_col, row) 
        for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
            for cell in cells:
                entry.append(cell.value)
                styles.append(type(cell.value))
        entries.append((entry,range,columns[index],styles))
    return entries


def work_with_line_in_structure_lessons(line,course_sheet, course_path):
    path=course_path
    sheet = course_sheet
    folder_col = get_column(sheet=sheet, row=1, name='Folders')
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    
    min_col, min_row,max_col,max_row=ranges[names.index('Level names')]
    level = []
    keywords = []
    keywords_cols = []
    for cells in sheet.iter_cols(min_col=min_col,min_row=2, max_col=max_col, max_row=2):
        for cell in cells:
            keywords.append(cell.value)
            keywords_cols.append(cell.column)
            print(cell.value)
    print(keywords)
    

    
    for cells in sheet.iter_cols(min_col=min_col,min_row=line, max_col=max_col, max_row=line):
        for cell in cells:
            level.append(cell.value)
            print(cell.value)
    
    dick= {}
  
    dick['MinimalPair_IPA'] = []
    dick['Second_MinimalPair_IPA'] = []
    dick['Target_IPA'] = []
    dick['Position'] = None
    dick['Extra'] = []
    extra=[]
    known = []
    #if id !=None:
    #    dick['_id'] = str(id)
    dick['L1-L2map_exception'] = None
    dick['Exercise_information'] = None
    dick["Exercise_name"] = None
    dick['Sidekick'] = [] 

        
    if level[keywords.index('Sound contrast')] != None:
        
        dick["Exercise_name"] =  level[keywords.index('Sound contrast')]
        if level[keywords.index('L1-L2map exception')]:

            dick['L1-L2map_exception'] = level[keywords.index('L1-L2map exception')]
        if level[keywords.index('Contrast information')]:

            dick['Exercise_information'] = level[keywords.index('Contrast information')]
        
        ipos = dick["Exercise_name"].split(' ')
    
        if len(ipos) > 1:
            dick['Position'] = ipos[-1]

    
        ipas = ipos[0].split('-')
        dick['Sidekick'] = ipas[-1]
        for ipa in ipas[0].split('/'):
            dick['MinimalPair_IPA'].append(ipa)
        for ipa in ipas[1].split('/'):
            dick['Second_MinimalPair_IPA'].append(ipa)
        

        
        #find group name in this case
      
        value = False
        current_row = line-1
        while not value:
            cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Target')])
            if cell.value:
                value = True
                print(cell.value, 'Group exercise')
                dick['Group_lesson'] = cell.value
            current_row -= 1


        
        
    
       
        for ipa in dick["Group_lesson"].split('/'):
            dick['Target_IPA'].append(ipa)



       
        
    elif level[keywords.index('Other contrast')] != None:

        dick["Exercise_name"] =  level[keywords.index('Other contrast')]
        dick['Sidekick'] = level[keywords.index('Other contrast')]
        if level[keywords.index('Contrast information')]:

            dick['Exercise_information'] = level[keywords.index('Contrast information')]
        value = False
        current_row = line-1
        while not value:
            cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Target')])
            if cell.value:
                value = True
                print(cell.value, 'Group exercise')
                dick['Group_lesson'] = cell.value
            current_row -= 1
    

    
    else:
        pass

   
    value = False
    current_row = line-1
    while not value:
        cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Lesson')])
        if cell.value:
            value = True
            print(cell.value, 'lesson')
            dick['Lesson'] = cell.value
        current_row -= 1

    value = False
    current_row = line-1
    dick['Level'] = None
    while not (value or current_row < 3):
        cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Course level')])
        if cell.value:
            value = True
            print(cell.value, 'level')
            dick['Level'] = cell.value
        current_row -= 1
   
    
    value = False
    current_row = line-1
    dick['Category'] = None
    while not (value or current_row < 3):
        cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Category')])
        if cell.value:
            value = True
            print(cell.value, 'level')
            dick['Category'] = str(cell.value)+'_'+str(current_row)
        current_row -= 1
    
    value = False
    current_row = line-1
    dick['Category_information'] = None
    while not (value or current_row < 3):
        cell = sheet.cell(current_row,column=keywords_cols[keywords.index('Category information')])
        
        
        if cell.value:
            value = True
        
            dick['Category_information'] = cell.value
        current_row -= 1
   
    #open an exercise.xlsx
    exercise_path = Path(sheet.cell(row=line, column=folder_col).value.replace('..',str(path.parent)))
    exercise_file = exercise_path/'exercise.xlsx'
    print(exercise_file.exists(), str(exercise_file))
    
    wb_exercise = load_workbook(str(exercise_file))

    data_row = 3
    sheet_ex = wb_exercise['Exercise']
    names,columns,ranges = get_sheet_structure(sheet = sheet_ex)
    c_prop = Counter(names)
    print(c_prop)
    
    for i in range(c_prop['Properties']):
            #it creates an entry with proper exercise id and id if it doesnt exist in db, excel file gets upodated, pay attention, 
            # exercises.xlsx shoul be closed
            
        entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_ex,table_name='Properties', names=names,ranges=ranges,row=data_row,columns=columns)[i]
        print(entry, entry[entry_columns.index('Key')])
        if entry[entry_columns.index('Key')] in known:
            if entry[entry_columns.index('Value')]:
                dick[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
            if entry[entry_columns.index('Description')]:
                dick[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
        else:
            if entry[entry_columns.index('Value')]:
                extra.append({str(entry[entry_columns.index('Key')]):str(entry[entry_columns.index('Value')]) })
            if entry[entry_columns.index('Description')]:
                extra.append({str(entry[entry_columns.index('Key')])+'_description':  str(entry_columns.index('Description'))})
    
    dick['Extra'] = extra
        

    
    case_vocab = False
    vocab = []
    case_mp = False
    mp = []
    case_nw = False
    nw = []
    print(wb_exercise.sheetnames, "mitya")
    
    
    for sheetname in wb_exercise.sheetnames:
        print(sheetname)
        if 'Vocab' in  sheetname:
            case_vocab = True
            vocab.append(sheetname)
        elif 'MP' in sheetname:
            #work on Vocab-Confusion Box
            case_mp = True
            mp.append(sheetname)
        elif 'NonWords' in sheetname:
            case_nw = True
            nw.append(sheetname)
        else:
            pass
  
    print(case_mp,case_vocab, case_nw)
    dick['MP_wordpairs']=[]
    dick['MP_nonwords'] = []
    
    if case_vocab:
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = False
        print(vocab)
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in vocab:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        for row in range(data_row,maximum_row+1):
            dick_word = {}
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
          
             #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            print(entry)
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            print(dick_word)
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
                
            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
     
            words.append(dick_word)

     
        dick['Vocabulary_words'] = words
        
        #with open('result.json', 'w') as fp:
        #    json.dump(dick, fp)
        
    if case_mp:

        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        
        vocab = mp
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in mp:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        switch = 0
        wordpair = []
        group_word = {}
        for row in range(data_row,maximum_row+1):
            if switch == 0:
                wordpair = []
            else:
                pass

            dick_word = {}
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
        
            #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet,table_name='ConfusionBoxes', names=names,ranges=ranges,row=row,columns=columns)[-1]
         
            
            bin = entry[-1]
            print(entry, bin, "bin")
            
            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            dick_word['bin'] = bin
            
            print(dick_word)
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]
                
            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
            wordpair.append(dick_word)
            if switch == 0:
                switch = 1
            else:
                switch = 0
                if str(bin) in group_word:
                    group_word[str(bin)].append(wordpair)
                else:
                    group_word[str(bin)] = []
                    group_word[str(bin)].append(wordpair)
                words.append(wordpair)

        
        
        dick['MP_wordpairs'] = words
        dick['Groups_of_pairs'] = group_word
                
   
        
     
    if case_nw:
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = False
        vocab = nw
        print(vocab)
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in vocab:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        
        sheet_cb = wb_exercise[sheet_name]
        names_cb,columns_cb,ranges_cb = get_sheet_structure(sheet = sheet_cb)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        words = []
        group_word = {}
        for row in range(data_row,maximum_row+1):
            dick_word = {}
            #start with confusionbox,

            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
          
             #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_cb,table_name='ConfusionBoxes', names=names_cb,ranges=ranges_cb,row=row,columns=columns_cb)[-1]
         
            
            bin = entry[-1]
            print(entry,bin,  "conf")
            
            
            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[-1]
            
            print(entry)
            
            dick_word['word'] = entry[entry_columns.index('Text')]
            dick_word['bin'] = bin
            print(dick_word)
            
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Properties', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[i]
                if entry[entry_columns.index('Key')]:
                    if entry[entry_columns.index('Value')]:
                        dick_word[entry[entry_columns.index('Key')]] = entry[entry_columns.index('Value')]
                    if entry[entry_columns.index('Description')]:
                        dick_word[entry[entry_columns.index('Key')]+'_description'] = entry[entry_columns.index('Description')]



            c = Counter(names_cb)
            print(c['Properties'], "here", names_cb)
            
            if c['Properties']==1:

                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_cb,table_name='Properties', names=names_cb,ranges=ranges_cb,row=row,columns=columns_cb)[-1]

                print(entry)
                if entry[entry_columns.index('Key')] == 'PairWord':
                    dick_word['nonword'] = entry[entry_columns.index('Value')]
                else:
                    pass
                
                
            
            c = Counter(names_wp)
            print(c['Properties'], "here")



            for num in [0,1]:
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                
                
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                audio = []
                for i in range(c_pron['Pronunciations']):
                    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Pronunciations', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[i]
                    print(entry, entry_columns)
                    

                    print(entry, entry[entry_columns.index('URI')])
                    if entry[entry_columns.index('URI')]:
                        audio.append(entry[entry_columns.index('URI')].replace('H:/Workspace/CalstFiles/WordObjectContent',''))
                dick_word['audio_'+str(num)] = audio    
     
            words.append(dick_word)
            if str(bin) in group_word:
                group_word[str(bin)].append(dick_word)
            else:
                group_word[str(bin)] = []
                group_word[str(bin)].append(dick_word)
            

        dick['MP_nonwords'] = words
        dick['Groups_of_singles'] = group_word
        
        #import json
        #with open('result.json', 'w') as fp:
        #    json.dump(dick, fp)
         
   

    else:
        pass

    return(dick)
 
               
def main(collection_name, course_folder):
    path = Path(course_folder)
    #the path to the course summary file
    structure_file  = Path(course_folder+'\\lessons_structure_to_json_JK.xlsx')

    if structure_file.exists():
        #check actions column for the command "submit"
        wb_structure = load_workbook(str(structure_file))
        sheet = wb_structure['Lessons']
        to_submit = []
        print(str(structure_file))
        action_col = get_column(sheet=sheet, row = 1, name='Actions')
        act_col = get_column(sheet=sheet, row = 1, name='Actions')
        level_1_key = get_column(sheet=sheet, row = 2, name='Category')
        level_2_key = get_column(sheet=sheet, row = 2, name='Lesson')
       
        level_1 = level_2 = ''
        for cells in sheet.iter_cols(min_col=action_col,min_row=3, max_col=action_col, max_row=sheet.max_row):
            
            for cell in cells:
                cell_action = sheet.cell(row=cell.row,column=act_col)
                if sheet.cell(row=cell.row, column=level_1_key).value:
                    level_1 = sheet.cell(row=cell.row, column=level_1_key).value
                if sheet.cell(row=cell.row, column=level_2_key).value:
                    level_2 = sheet.cell(row=cell.row, column=level_2_key).value
              
                #if str(cell.value).lower() == 'retract':
                if cell.value == 'submitted':
                    to_submit.append((level_1,level_2, cell.row ))
                    print(cell.value)
        # list to_submit contains rows of summary file to submit
        print('rows to_submit: ', to_submit)
        
        
        
        
        
        
        
        for line in to_submit:
            cell_action = sheet.cell(row=line[-1],column=act_col)
            #try:
            dictionary = work_with_line_in_structure_lessons(line=line[-1],course_sheet=sheet,course_path=path)
            print(dictionary, "dic")
            
            
            
            dictionary['Excel_weight'] = line[-1]
            cell_action.value = 'resubmitted'
            wb_structure.save(structure_file)

            collection_name.insert_one(dictionary)
            
                        #except:
                #cell_action.value = 'failed_submitted'
                #wb_structure.save(structure_file)  
                     
        
    else:
        print("No exercise file here. quitting")
        exit() 

if __name__ == "__main__":
   
    language='Greek'
    dbname = get_database('calst_new_system_bin')
    collection_name = dbname[language]
    #collection_name = dbname['testlang']
    output_sound = r'C:\Source\Repos\CalstEnglish\CalstFiles\WordObjectContent'+'\\'+language+r'\OriginalWords_Wav'
    print(output_sound)
 
    keyword = 'retract'	
    dst_path = Path(output_sound)
    #folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\'
    #folder = 'G:\\My Drive\\CALST_courses\\'+str(language)+'_course_styled\\'
    #folder = 'C:\\Source\\Repos\\mysql-excel\\'+str(language)+'_course_styled\\'
    folder = r'C:\Users\dmitrysh\OneDrive - NTNU\CALST_courses\\'+str(language)+'_course_styled'
    #folder = r'C:\Users\dmitrysh\OneDrive - NTNU'
    main(collection_name = collection_name, course_folder=folder)
    #cnxn.commit()

    #cnxn.close()