#to be used from actions_templates
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import cols_from_range, coordinate_to_tuple,get_column_interval

def get_sheet_structure(sheet):
    #a database row specified as merged cells
    merged_ranges = sheet.merged_cells.ranges
    database = []
    ranges = []
    for range in merged_ranges:
        (col_low, row_low, col_high, row_high) = range.bounds
        ranges.append(range.bounds)
        database.append(sheet.cell(row_low,col_low).value)
    database_and_ranges = zip(database, ranges)
    database_cols = []
    for data in database_and_ranges:
        (col_low, row_low, col_high, row_high) = data[-1]
        if row_high-row_low != 0:
            print("Error,too many rows for info caption")
            exit()
        else:
            pass
        
        col_id = []
        for row in sheet.iter_rows(min_col=col_low,max_col=col_high,min_row=row_low+1,max_row=row_high+1):
            for cell in row:
                col_id.append(cell.value)
        database_cols.append(col_id)

    return database, database_cols,ranges


def template_out_of_exercise(folder):
    folderout = folder.replace('C:\Source\Repos\mysql-excel','G:\My Drive\CALST_courses')
    folderout_path = Path(folderout)
    folder_path = Path(folder)
    excel = folder_path/'exercise.xlsx'
    excel_out = folderout_path/'exercise_template.xlsx'
    if excel.exists():
        print(str(excel), " exists")
    wb = load_workbook(filename = excel)
    for i in wb.sheetnames:
        print(i)
        sheet = wb[i]

        for col in range(1,sheet.max_column+1):
            for row in range(3,sheet.max_row+1):
                sheet.cell(column=col, row=row).value = None
    
        
    wb.save(str(excel_out))
    if excel_out.exists():
        print(str(excel_out), " exists")
def main(folder):
    
    
    
   


    
    template_out_of_exercise(folder = folder)

    

   
   

if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='python diff_folder_and_mysql.py -f foldername')
    parser.add_argument('-f',dest='folder')
    parser.add_argument('-diff',dest='diff',default='diff.txt')
    args = parser.parse_args()
    if args.folder:
        main(folder = args.folder, diff=args.diff)
    else:
        #folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\The alphabet'
        folder = r'C:\Source\Repos\mysql-excel\Italian_course_styled'
        main(folder=folder)