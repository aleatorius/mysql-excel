from pathlib import Path
from unittest.util import _MIN_COMMON_LEN
from openpyxl import load_workbook


def get_sheet_structure(sheet):
    #a database row specified as merged cells
    merged_ranges = sheet.merged_cells.ranges
    database = []
    ranges = []
    for range in merged_ranges:
        (col_low, row_low, col_high, row_high) = range.bounds
        ranges.append(range.bounds)
        database.append(sheet.cell(row_low,col_low).value)
    database_and_ranges = zip(database, ranges)
    database_cols = []
    for data in database_and_ranges:
        (col_low, row_low, col_high, row_high) = data[-1]
        if row_high-row_low != 0:
            print("Error,too many rows for info caption")
            exit()
        else:
            pass
        
        col_id = []
        for row in sheet.iter_rows(min_col=col_low,max_col=col_high,min_row=row_low+1,max_row=row_high+1):
            for cell in row:
                col_id.append(cell.value)
        database_cols.append(col_id)

    return database, database_cols,ranges


def main(folder):

    #main code, can be referenced from another codes as well

    path = Path(folder)
    exercise_path  = path/'exercise_template.xlsx'
    if exercise_path.exists():
        print('In the folder: ', str(path), ' exercise.xlsx exists, processing it')
        wb = load_workbook(str(exercise_path))
        #let's find all sheets
        sheets = []
        for i in wb.sheetnames:
            sheets.append(i)
        print(sheets)
        #we get a list like: ['Exercise', 'Vocab- ConfusionBoxes', 'Vocab- Words Properties', 'Vocab- Speaker_Trans_0', 'Vocab- Speaker_Trans_1']
        #let's work with 'Vocab- ConfusionBoxes'
        
        sheet = wb['Vocab- ConfusionBoxes']
        
        #let us request the structure of this sheet, for this I wrote the function get_sheet_structure
        names_of_tables, table_cols,ranges = get_sheet_structure(sheet=sheet)
        
        print(names_of_tables)
        #['ConfusionBoxes', 'TranscriptionConfusionBoxes', 'Words', 'Transcriptions', 'Transcriptions']
        
        print(table_cols)
        #this is a list of lists:
        #[['Id', 'ExerciseId', 'DialectId', 'CorrectTranscriptionId', 'Bin'], ['Transcription_Id', 'ConfusionBox_Id'], ['Id', 'Text', 'LanguageId'], ['Id', 'WordId', 'Text'], ['Id', 'WordId', 'Text']]
        # to confusionboxes corresponds the list ['Id', 'ExerciseId', 'DialectId', 'CorrectTranscriptionId', 'Bin']
        #to Words -['Id', 'Text', 'LanguageId']
        #and we have two transcriptions tables -- for each speaker, MLHH and ARS
        
        print(ranges)
        #this gives the column-row ranges for table_names which are merged cells
        #[(1, 1, 5, 1), (6, 1, 7, 1), (8, 1, 10, 1), (11, 1, 13, 1), (14, 1, 16, 1)]
        
        #for example we are interested in Words
        words_range = ranges[2]
        print(words_range)
        #(8, 1, 10, 1)
        #let's define max and min col and row for this range
        col_min, row_min,col_max,row_max = words_range
        #let us iterate over this block
        entries = []
        for cells in sheet.iter_rows(min_col=col_min,min_row=row_min, max_col=col_max,max_row=sheet.max_row):
            entry = []
            for cell in cells:
                entry.append(cell.value)
            entries.append(entry)
        print(entries)
        #it is not that full now, only first two rows are not empty -- 
        #[['Words', None, None], ['Id', 'Text', 'LanguageId'], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None], [None, None, None]]
        # take into account that merged cells are printed like this ['Words', None, None]

        #ok, we want to add the following word list together with its language ids 
     
        word_list = [('boka',5),('poka',5),('sembra',5),('sempra',5)]
        #that's into position *text* and *languageid*

        text_column = col_min+ table_cols[names_of_tables.index('Words')].index('Text')
        languageid_column = col_min + table_cols[names_of_tables.index('Words')].index('LanguageId')
        for row,word in enumerate(word_list):
            sheet.cell(row=row+3,column=text_column).value = word[0]
            sheet.cell(row=row+3,column=languageid_column).value = word[1]
        wb.save(str(path/'exercise_new.xlsx'))
        #check the file





        
if __name__ == "__main__":
    folder = r'G:\My Drive\CALST_courses\Spanish_course_styled\Beginner\Lesson 9\Math'
    main(folder=folder)