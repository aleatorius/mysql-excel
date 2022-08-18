from re import A
from textwrap import indent
from natsort import natsorted
from itertools import groupby
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import diff_folder_and_mysql
import course_structure_db_excel


def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()


def main(folder):
    path = Path(folder)
    print(path.parent)
    structure_path  = Path(folder+'\\lessons_structure.xlsx')

    if structure_path.exists():
        wb_s = load_workbook(str(structure_path))
        sheet = wb_s.active
        to_diff = []
        to_submit = []
        action_col = get_column(sheet=sheet,row=1,name='Actions')
        folder_col = get_column(sheet=sheet,row=1,name='Folders')
        
        for cells in sheet.iter_cols(min_col=action_col,min_row=3, max_col=action_col, max_row=sheet.max_row):
            for cell in cells:
                if cell.value:
                    if cell.value.lower() == 'diff':
                        cell_f = sheet.cell(column=folder_col, row=cell.row)
                        path_folder = str(path.parent)+cell_f.value.replace('..','')
                        to_diff.append(path_folder)
                    if cell.value.lower() == 'submit':
                        cell_f = sheet.cell(column=folder_col, row=cell.row)
                        path_folder = str(path.parent)+cell_f.value.replace('..','')
                        to_submit.append(path_folder)
        if to_diff:
            for todiff in to_diff:
                print(todiff)
                diff_folder_and_mysql.main(folder=todiff, output_diff='test.txt')
        if to_submit:
            for submit in to_submit:
                print(submit)
    else:
        print(str(structure_path.parent))
        print(str(structure_path.name))
        
        course_structure_db_excel.main(folder=str(structure_path.parent), output=str(structure_path.name ))
        print("No file with lesson info, have created one. You can add actions here: "+ str(structure_path))
    
  
if __name__ == "__main__":

    #file = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\lessons_structure.xlsx'
    folder = 'G:\\My Drive\\CALST_courses\\Spanish_course_styled\\'
    main(folder=folder)