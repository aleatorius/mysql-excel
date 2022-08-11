import pyodbc 
import argparse
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils.cell import cols_from_range, coordinate_to_tuple,get_column_interval

def compare(entry, data):
    if len(entry) == len(data):
        diff = []
        for ind,ex in enumerate(entry):
            if ex == data[ind]:
                diff.append(True)
            else:
                diff.append(False)
        return diff
    else:
        print("ERROR in compare entries", entry, data)
            
def get_full_request(cursor, database, data, entry):
    sqlcom = 'SELECT * FROM ['+database+'].[dbo].[' + data[0]+'] WHERE '
    count = 0
    for ind, ex in enumerate(data[1]):
        if count < len(entry)-1:
            sqlcom = sqlcom + str(ex) + ' = ' + str(entry[ind]) + ' AND '
        else:
            sqlcom = sqlcom + str(ex) + ' = ' + str(entry[ind]) 
        count = count + 1 
    cursor.execute(sqlcom)
    output = [list(i) for i in cursor.fetchall()] 
    return output    

def get_sheet_structure(sheet):
    #a database row specified as merged cells
    merged_ranges = sheet.merged_cells.ranges
    database = []
    ranges = []
    for range in merged_ranges:
        (col_low, row_low, col_high, row_high) = range.bounds
        ranges.append(range.bounds)
        database.append(sheet.cell(row_low,col_low).value)
    database_and_ranges = zip(database, ranges)
    database_cols = []
    for data in database_and_ranges:
        (col_low, row_low, col_high, row_high) = data[-1]
        if row_high-row_low != 0:
            print("Error,too many rows for info caption")
            exit()
        else:
            pass
        
        col_id = []
        for row in sheet.iter_rows(min_col=col_low,max_col=col_high,min_row=row_low+1,max_row=row_high+1):
            for cell in row:
                col_id.append(cell.value)
        database_cols.append(col_id)

    return database, database_cols,ranges

def diff_write(diff,diff_file,output,entry,data,sheet,row):
    indices = [i for i, x in enumerate(diff) if x == False]
    (min_col, min_row, max_col, max_row) = data[-1]
    temp = get_column_interval(min_col, max_col)
    for index in indices:
        if not output[index]:
            print(output[index],"is empty")
        else:
            pass
        diff_file.write('Sheet: "'+sheet.title+ '" Cell:"'+ str(temp[index])+str(min_row+row)+' ' + str(data[1][index]) +'"\n Excel: "'+ str(entry[index]) +'" db: "'+ str(output[index])+'"\n\n')

def noentry_write(noentry_file,entry,data,sheet,row):
    (min_col, min_row, max_col, max_row) = data[-1]
    temp = get_column_interval(min_col, max_col)
    noentry_file.write('NO ENTRY! Sheet: "'+sheet.title+ '" Cell:"'+ str(temp)+str(min_row+row)+' ' + str(data[1]) +'"\n Excel: "'+ str(entry) +'\n\n')

def excel_in_folder(excel,cursor,database,output_diff):

    istherediffoutput = False
    firstrun = True
    
    wb = load_workbook(filename = excel)
    for i in wb.sheetnames:
        print(i)
        sheet = wb[i]
        names,columns,ranges = get_sheet_structure(sheet = sheet)
        
        for row in range(2,sheet.max_row):
            for data in zip(names,columns,ranges):
                (col_low, row_low, col_high, row_high) = data[-1]
                entry = []
                for cells in sheet.iter_cols(min_col=col_low,min_row=row_low+row, max_col=col_high, max_row= row_low+row):
                    for cell in cells:
                        entry.append(cell.value)
                
                if all(x is None for x in entry):
                    pass
                else:
                    if data[0] == 'WrapperExercises':
                        #it should be defined by both columns
                        output = get_full_request(database=database,cursor=cursor,data=data, entry=entry)
                        if len(output) == 0:
                            
                            if firstrun == True:
                                if os.path.isfile(output_diff):
                                    diff_file = open(output_diff,'a')
                                else:
                                    diff_file = open(output_diff,'w')
                                firstrun = False
                                diff_file.write('Data file :'+str(excel)+'\n\n')      
                            else:
                                pass
                            istherediffoutput = True
                            noentry_write(noentry_file=diff_file,entry=entry,data=data,sheet=sheet, row=row)
                        elif len(output)>1:
                            print("ERROR, not unique entry, exiting ", data[0] )  
                            exit()
                        else:
                            pass
                    elif data[0] == 'TranscriptionConfusionBoxes':
                        output = get_full_request(database=database,cursor=cursor,data=data, entry=entry)
                        if len(output) ==0:
                            if firstrun == True:
                                if os.path.isfile(output_diff):
                                    diff_file = open(output_diff,'a')
                                else:
                                    diff_file = open(output_diff,'w')
                                firstrun = False
                                diff_file.write('Data file :'+str(excel)+'\n\n')
                            else:
                                pass
                            istherediffoutput = True
                            noentry_write(noentry_file=diff_file,entry=entry,data=data,sheet=sheet, row=row)

                        elif len(output)>1:
                            print("ERROR, not unique entry, exiting ", data[0] )  
                            exit()
                        else:
                            pass
                    else:
                        sqlcom = 'SELECT * FROM ['+database+'].[dbo].[' + data[0]+'] WHERE '
                        sqlcom = sqlcom + str(data[1][0]) + ' = ' + str(entry[0])
                        cursor.execute(sqlcom)
                        output = [list(i) for i in cursor.fetchall()] 
                        if len(output) == 1:
                            diff = compare(entry=entry,data=output[0] )
                            if False in diff:
                                istherediffoutput = True
                                if firstrun == True:
                                    if os.path.isfile(output_diff):
                                        diff_file = open(output_diff,'a')
                                    else:
                                        diff_file = open(output_diff,'w')
                                    firstrun = False
                                    diff_file.write('Data file :'+str(excel)+'\n\n')
                                else:
                                    pass
                                diff_write(diff=diff,diff_file=diff_file,output=output[0],entry=entry,data=data,sheet=sheet, row=row)
                        elif len(output) == 0:
                            if firstrun == True:
                                if os.path.isfile(output_diff):
                                    diff_file = open(output_diff,'a')
                                else:
                                    diff_file = open(output_diff,'w')
                                firstrun = False
                                diff_file.write('Data file :'+str(excel)+'\n\n')
                                    
                            else:
                                pass
                            istherediffoutput = True
                            noentry_write(noentry_file=diff_file,entry=entry,data=data,sheet=sheet, row=row)
                        else:
                            print('wtf', entry, output)
                            exit()
    if istherediffoutput == True:
        diff_file.close
    

def main(folder,output_diff):
    folder_path = Path(folder)
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    ser_file.close
    #connect to the calst database
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    
    folder_path = Path(folder)
    if folder_path.exists():
        pass
    else:
        print(' There is no such folder as "'+folder+'"','\n Please enter the correct folder name')
        exit()

    excels = list(folder_path.glob('*.xlsx'))
    for excel in excels:
        if '~$' in str(excel):
            pass
        elif 'wrapper.xlsx' in str(excel):
            excel_in_folder(cursor=cursor,database=database,output_diff=output_diff,excel=excel)
        elif 'exercise.xlsx' in str(excel):
            excel_in_folder(cursor=cursor,database=database,output_diff=output_diff,excel=excel)
        else:
            pass

   
   

if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='python diff_folder_and_mysql.py -f foldername')
    parser.add_argument('-f',dest='folder')
    parser.add_argument('-diff',dest='diff',default='diff.txt')
    args = parser.parse_args()
    if args.folder:
        main(folder = args.folder, diff=args.diff)
    else:
        Ъfolder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\The alphabet'
        folder = 'C:\Source\Repos\mysql-excel\Spanish_course\Beginner\Lesson 1'
        main(folder=folder, output_diff='difft.txt')