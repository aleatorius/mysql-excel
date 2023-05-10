import pyodbc 
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_header(cursor,table):
    sqlcom = "SELECT * FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'"+table+"'"
    cursor.execute(sqlcom)  
    data = list(cursor.fetchall())
    header = [i[3] for i in data]
    header_0 = len(data)*[table]
    return header_0,header

def get_zipped(data):
    zipped = []
    for i in data:
        line=[]
        for j in i:
            line=line+j
        zipped.append(line)
    return zipped

def get_confusionboxpage(cursor, wbout, input_data,title):
    data_cb = []
    for i in input_data:
        sqlcom = "SELECT * FROM [dbo].[ConfusionBoxes] WHERE Id = "+str(i[1])
        cursor.execute(sqlcom)
        data_cb.append([list(i) for i in cursor.fetchall()][0])
    
    #data_coupled, data_cb -- transcriptions + confusionbox info
    data_0 = input_data        

    data_nonword = []
    for i in data_cb:
        sqlcom = "SELECT * FROM [dbo].[Properties] WHERE [Key] = 'PairWord' AND  Description = "+str(i[0])
        cursor.execute(sqlcom)
        
        ls = [list(i) for i in cursor.fetchall()]
        if ls:
            data_nonword.append(ls[0])
        else:
            data_nonword.append(ls)    

    lengths_nonword = [len(i) for i in data_nonword]

    #iterate over transcriptions ids
    multi=[]
    lengths = []
    for en in data_0:
        sqlcom = "SELECT * FROM [dbo].[Transcriptions] WHERE Id = "+str(en[0])
        cursor.execute(sqlcom)
        trans = [list(i) for i in cursor.fetchall()][0]
        sqlcom = "SELECT * FROM [dbo].[Transcriptions] WHERE WordId = "+str(trans[1])
        cursor.execute(sqlcom)
        trans = [list(i) for i in cursor.fetchall()]
        multi.append(trans)
        lengths.append(len(trans))
    multi_zipped = get_zipped(multi)

    data_word=[]
    for en in multi:
        sqlcom = "SELECT * FROM [dbo].[Words] WHERE Id = "+str(en[0][1])
        cursor.execute(sqlcom)
        data_word.append([list(i) for i in cursor.fetchall()][0])
    
    header_cells = []
    sheet = wbout.create_sheet(title+'- ConfusionBoxes')
    header_0,header = get_header(cursor=cursor,table='ConfusionBoxes')
    header_cells.append(len(header))
    header_temp_0,header_temp = get_header(cursor=cursor,table='TranscriptionConfusionBoxes')
    header_cells.append(len(header_temp))
    header_0 = header_0 + header_temp_0
    header = header + header_temp
    header_temp_0,header_temp = get_header(cursor=cursor,table='Words')
    header_cells.append(len(header_temp))
    header_0 = header_0 + header_temp_0
    header = header + header_temp
    header_temp_0,header_temp = get_header(cursor=cursor, table='Transcriptions')
    header_0 = header_0 + max(lengths)*header_temp_0
    header = header + max(lengths)*header_temp
    for i in range(max(lengths)):
        header_cells.append(len(header_temp))
    
    # depending ion existing nonwords
    
    if max(lengths_nonword) == 0:
        megadata = zip(data_cb, input_data, data_word,multi_zipped)
        matrix = get_zipped(megadata)
    else:
        megadata = zip(data_cb, input_data, data_word,multi_zipped,data_nonword)
        matrix = get_zipped(megadata)
        header_temp_0,header_temp = get_header(cursor=cursor,table='Properties')
        header_0 = header_0 + header_temp_0
        header = header + header_temp
        header_cells.append(len(header_temp))
    sheet.append(header_0)
    sheet.append(header)
    for i in matrix:
        sheet.append(i)     
    
    get_style(header_cells,sheet)
    return data_word,multi

def get_style(header_cells,sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            if cell.row > 1:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))+2.7
                except:
                    pass
            else:
                pass
        adjusted_width = (max_length) 
        sheet.column_dimensions[column].width = adjusted_width   
    start=1
    twocolor = ['00CCFFCC','00FFFF99','00C0C0C0','0033CCCC']
    ccolor = twocolor[0]
    twopattern = ['lightDown','darkGray']
    for i in header_cells:
        cell_header = sheet.cell(row=1, column=start)
        double = Side(border_style="double", color="00008000")
        cell_header.border = Border(top=double, left=double, right=double, bottom=double)
        cell_header.font  = Font(b=True, color="00008000", size = 8)
        cell_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False,  shrink_to_fit=True)
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start+i-1)
        for j in range(i):
            cell_h = sheet.cell(row=2, column=start+j)
            cell_h.font  = Font(b=True, color="00008000", size = 10)
            cell_h.alignment = Alignment(horizontal="general", vertical="bottom", wrap_text=False,  shrink_to_fit=True)
            if j == 0:
                cell_h.border = Border(left=double, bottom=double)
            else:
                cell_h.border = Border(bottom=double)

        
        ppattern = twopattern[0]
        for l in range(3,sheet.max_row+1):
            for j in range(i):
                cell_ordinary = sheet.cell(row = l,column=start+j)
                cell_ordinary.fill = PatternFill(ppattern, fgColor=ccolor)
                thin = Side(border_style="thin", color="000000")
                cell_ordinary.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if ppattern == twopattern[0]:
                ppattern = twopattern[1]
            else:
                ppattern = twopattern[0]
        
        start = start + i
        if ccolor == twocolor[0]:
            ccolor = twocolor[1]
        else:
            ccolor = twocolor[0]
    
    if header_cells[-1] == 1:
        cell_h.border = Border(bottom=double, right=double,left=double)   
    else:
        cell_h.border = Border(bottom=double, right=double)     

def get_wordproperties(cursor,wbout,input_data,title):
    header_cells = []
    header_0,header=  get_header(cursor=cursor, table='Words')
    header_cells.append(len(header))
    header_0_ex,header_ex =  get_header(cursor=cursor, table='Properties')
    sheet = wbout.create_sheet(title+'- Words Properties')
    
    temp = []
    for  en in input_data:
        sqlcom = "SELECT * FROM [dbo].[Properties] WHERE ExerciseId IS NULL AND WordId = "+str(en[0])
        cursor.execute(sqlcom)
        ls=[list(i) for i in cursor.fetchall()]
        temp.append(ls)
    zipped_word=[]
    lengths_word = []
    for i in temp:
        l = []
        lengths_word.append(len(i))
        for j in i:
            l = l + j
        zipped_word.append(l)
    temp = zipped_word
    
    header_0 = header_0 + max(lengths_word)*header_0_ex
    header = header + max(lengths_word)*header_ex
    
    for i in range(max(lengths_word)):
        header_cells.append(len(header_ex))

    megadata = zip(input_data,temp)
    matrix=[]
    for i in megadata:
        line=[]
        for j in i:
            line=line+j
        matrix.append(line)

    sheet.append(header_0)
    sheet.append(header)
    for i in matrix:
        sheet.append(i)    

    get_style(header_cells,sheet)

def get_transcription(cursor, wbout, i , word_data, multi, title):
    sheet = wbout.create_sheet(title+'- Speaker_Trans_'+str(i))
    header_cells = []
    header_0,header = get_header(cursor=cursor, table='Words')
    header_cells.append(len(header))
    header_temp_0,header_temp = get_header(cursor=cursor, table='Transcriptions')
    header_0 = header_0 + header_temp_0
    header = header + header_temp
    header_cells.append(len(header_temp))
    header_temp_0,header_temp  =  get_header(cursor=cursor, table='Pronunciations')
   
    data_3 = []
    for  en in multi:
        ls = []
        try:
            sqlcom = "SELECT * FROM [dbo].[Pronunciations] WHERE Transcription_Id = "+str(en[i][0])
            cursor.execute(sqlcom)
            ls.append([list(i) for i in cursor.fetchall()])
        except:
            pass
        data_3.append(ls)
    
    zipped=[]
    lengths_prop = []
    for en in data_3:
        temp = []
        for entry in en: 
            l = []
            lengths_prop.append(len(entry))
            for j in entry:
                l = l + j
            temp.append(l)
        zipped.append(temp)
    
    header_0 = header_0 + max(lengths_prop)*header_temp_0
    header = header + max(lengths_prop)*header_temp
    for l in range(max(lengths_prop)):
        header_cells.append(len(header_temp))
    
    sheet.append(header_0)
    sheet.append(header)

    for j in zip(word_data,multi,zipped):
        line = j[0]
        try:
            line = line + j[1][i]+j[2][0]
        except:
            pass
        sheet.append(line)
    get_style(header_cells,sheet)

def wrapper(cursor,data,folder):
    wbout = Workbook()
    header_cells = []
    header_0,header = get_header(cursor=cursor,table='Wrappers')
    header_cells.append(len(header))
    sheet = wbout.active
    sheet.title = "Wrappers"
    header_0,header =  get_header(cursor=cursor, table='Wrappers')
    sheet.append(header_0)
    sheet.append(header)
    sheet.append(data)
    get_style(header_cells,sheet)
    wbout.save(folder+'\wrapper.xlsx')
    wbout.close()

def get_wrapper_file(cursor,input_folder,input_wrapper):
    wbout = Workbook()
    header_cells = []
    header_0,header = get_header(cursor=cursor, table='Wrappers')
    header_cells.append(len(header))
    sheet = wbout.active
    sheet.title = "Wrappers"
    sqlcom = "SELECT * FROM [dbo].[Wrappers] WHERE Id = "+input_wrapper
    cursor.execute(sqlcom)
    data = [list(i) for i in cursor.fetchall()]
    
    #if there are properties for wrappers
    #the third block, Pronunciation, can be several properties for an exercises

    header_0_ex,header_ex =  get_header(cursor=cursor, table='Properties')
    sqlcom = "SELECT * FROM [dbo].[Properties] WHERE ExerciseId IS NULL AND WrapperId = "+input_wrapper
    cursor.execute(sqlcom)
    data_1=[list(i) for i in cursor.fetchall()]

    zipped=[]
    lengths = [len(data_1)]
    for i in data_1:
        zipped = zipped + i
    data_1 = zipped
    data_combined = data[0] + data_1
    
    header_0 = header_0 + max(lengths)*header_0_ex
    header = header + max(lengths)*header_ex
    for i in range(max(lengths)):
        header_cells.append(len(header_ex))
    sheet.append(header_0)
    sheet.append(header)
    sheet.append(data_combined)
    get_style(header_cells,sheet)
    wbout.save(input_folder+"\wrapper.xlsx")
    wbout.close()

def get_exercise_infopage(cursor,wbout,entry,input_wrapper):
    header_cells = []
    header_0,header = get_header(cursor=cursor, table='WrapperExercises')
    header_cells.append(len(header))
    sqlcom = "SELECT * FROM [dbo].[WrapperExercises] WHERE Exercise_Id = "+str(entry) + " AND Wrapper_Id = "+input_wrapper
    cursor.execute(sqlcom)
    data_wrapper = [list(i) for i in cursor.fetchall()][0]
    sheet = wbout.active
    header_0_ex,header_ex =  get_header(cursor=cursor, table='Exercises')
    header_0 = header_0 + header_0_ex
    header = header + header_ex
    header_cells.append(len(header_ex))
    sheet.title = "Exercise"

    sqlcom = "SELECT * FROM [dbo].[Exercises] WHERE Id = "+str(entry)
    cursor.execute(sqlcom)
    data_exercises=[list(i) for i in cursor.fetchall()][0]
    sqlcom = "SELECT * FROM [dbo].[Properties] WHERE ExerciseId = "+str(entry)
    cursor.execute(sqlcom)
    data_properties=[list(i) for i in cursor.fetchall()]
    
    item = "ExerciseName"
    length = len(data_properties)
    
    if length == 1:
        data_properties = data_properties[0]
    else:
        zipped=[]
        length = len(data_properties)
        for i in data_properties:
            zipped = zipped + i                
        data_properties = zipped
    
    if "MinimalPair" in data_properties:
            print("it is a minimal pair exercise")
            MinimalPair = True
    else:
        MinimalPair = False
    
    if data_properties.index(item):
        folder_exercise = data_properties[data_properties.index(item)-1].replace(":","_colon").replace("?","_qm_").replace('/','_backslash_')
    else:
        print("error, no name of exercises in properties")
    
    header_0_ex,header_ex =  get_header(cursor=cursor,table='Properties')
    header_0 = header_0 + length*header_0_ex
    header = header + length*header_ex
    for i in range(length):
        header_cells.append(len(header_ex))

    #combine all blocks
    data_combined = data_wrapper + data_exercises + data_properties
    sheet.append(header_0)
    sheet.append(header)
    sheet.append(data_combined)
   
    #styling
    get_style(header_cells,sheet)
    
    #minimalpair sheet may be redundant
    if MinimalPair == True:
        #create extra sheet
        sqlcom = "SELECT * FROM [dbo].[Properties] WHERE WordID IS NOT NULL AND ExerciseId = "+str(entry)
        cursor.execute(sqlcom)
        data_word_mp =[list(i) for i in cursor.fetchall()]
        
        if data_word_mp:
            sheet = wbout.create_sheet('MinimalPair_Word_LIst')
            header_0,header = get_header(cursor=cursor,table='Properties')
            header_word_0,header_word = get_header(cursor=cursor, table='Words')
            sheet.append(header_word_0+header_0)  
            sheet.append(header_word+header)
            for i in data_word_mp:
                sqlcom = "SELECT * FROM [dbo].[Words] WHERE Id = "+str(i[3])
                cursor.execute(sqlcom)
                word = [list(i) for i in cursor.fetchall()][0]
                sheet.append(word+i)
        else:
            pass    
    return folder_exercise

def wrapper_to_folder(cursor,input_folder,input_wrapper):
    print(input_wrapper)

    #excel file with wrapper info
    get_wrapper_file(cursor=cursor, input_folder=input_folder,input_wrapper=input_wrapper)

    #list of exercises

    sqlcom = "SELECT * FROM [CalstContent].[dbo].[WrapperExercises] WHERE Wrapper_Id = "+input_wrapper
    cursor.execute(sqlcom)
    data = [list(i) for i in cursor.fetchall()]
    print(data)
    
    exercises = [i[1] for i in data]
    print(exercises)
    
    #Exercises, folders
    for entry in exercises:
        wbout = Workbook()
        #the first sheet, with wrapper info, from wrappers db
        folder_exercise = get_exercise_infopage(cursor=cursor,wbout=wbout, entry=entry, input_wrapper=input_wrapper)
        folder_exercise = input_folder+'/'+folder_exercise

        result = True
        while result is True:
            p = Path(folder_exercise)
            result = p.exists()
            print(result, folder_exercise)
            if result == False:
                pass
            else:
                folder_exercise = folder_exercise + '_extra'
        print(result, folder_exercise)   
        Path(folder_exercise).mkdir(parents=True, exist_ok=False)
        
        print(folder_exercise)
        #create sheet confusionbox and transcriptions
        #get the list of confusion box numbers, pay attention, for MP exercise each confusionbox corresponds to either minimal pair words, 
        # or to just one word. for vocabulary one confusion box can correspond to all the words for this exercise
        sqlcom = "SELECT * FROM [dbo].[ConfusionBoxes] WHERE ExerciseId = "+str(entry)
        cursor.execute(sqlcom)
        data = list(cursor.fetchall())
        data_confusion=[list(i) for i in data]
        
        data_vocab = []
        data_nonword = []
        data_pair = []
        data_threesyll = []

        for i in data_confusion:
            sqlcom = "SELECT * FROM [dbo].[TranscriptionConfusionBoxes] WHERE ConfusionBox_Id = "+str(i[0])
            cursor.execute(sqlcom)
            temp = [list(i) for i in cursor.fetchall()]
            if len(temp) == 2:
                data_pair.append(temp)
            elif len(temp) == 1:
                data_nonword.append(temp)
            elif len(temp) == 3:
                data_threesyll.append(temp)
            elif len(temp) > 3:
                data_vocab.append(temp)
            else:
                pass
            
        alldata = [data_pair, data_nonword, data_vocab, data_threesyll]
        alltitle = ['MP', 'NonWords','Vocab',"ThreeSyll"]
        data_and_title = zip(alldata,alltitle)

        for en in data_and_title:
            if en[0]:
                ex = en[1]
                data_coupled = []
                for i in en[0]:
                    for j in i:
                        data_coupled.append(j)
                word_list, multi = get_confusionboxpage(cursor=cursor, wbout=wbout, input_data=data_coupled,title=ex) 
                get_wordproperties(cursor,wbout,word_list, title=ex)
                for i in range(max([len(i) for i in multi])):
                    get_transcription(cursor=cursor,wbout=wbout, i=i, word_data=word_list, multi=multi,title=ex)
        wbout.save(folder_exercise+'/exercise.xlsx')
        wbout.close()


def main():
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info

    #connect to the calst database

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()

    #wrapper corresponds to a second levels structures like lessons 1, 2,3 

    localcase = False

    if localcase:
        Path('Test_nor').mkdir(parents=True, exist_ok=True)
        #wrapper_to_folder(cursor,"Test","1296")
        #wrapper_to_folder(cursor,"Test","1302")
        wrapper_to_folder(cursor,"Test","79")
        #wrapper_to_folder(cursor,"Easy/Italian","1293")
    else:  
        course = 'English'
        sqlcom = "SELECT * FROM [CalstContent].[dbo].[Wrappers] where [Name] = '"+course+"EntryPoint'"
        cursor.execute(sqlcom)  
        data = [list(i) for i in cursor.fetchall()] 
        folder_level_0 = course+'_course_revised'
        Path(folder_level_0).mkdir(parents=True, exist_ok=True)
        wrapper(cursor, data[0], folder_level_0)
        sqlcom = "SELECT * FROM [dbo].[Wrappers] WHERE WrapperId = "+ str(data[0][0])
        cursor.execute(sqlcom)  
        data = [list(i) for i in cursor.fetchall()]
        for i in data:
            folder_level_1 = folder_level_0+'/'+str(i[2])
            Path(folder_level_1).mkdir(parents=True, exist_ok=True)
            wrapper(cursor,i,folder_level_1)
            sqlcom = "SELECT * FROM [dbo].[Wrappers] WHERE WrapperId = "+ str(i[0])
            cursor.execute(sqlcom)  
            ls = [list(i) for i in cursor.fetchall()] 
            
            for j in ls:
                folder_level_2 = folder_level_1+'/'+str(j[2])
                Path(folder_level_2).mkdir(parents=True, exist_ok=True)
                print(folder_level_2)
                wrapper_to_folder(cursor,folder_level_2,str(j[0]))
                sqlcom = "SELECT * FROM [dbo].[Wrappers] WHERE WrapperId = "+ str(j[0])
                cursor.execute(sqlcom)  
                level_3 = [list(i) for i in cursor.fetchall()] 

                for entry in level_3:          
                    folder_level_3 = folder_level_2+'/'+str(entry[2]).replace(":","_colon").replace("?","_qm_").replace('/','_backslash_').replace("'\'r'\'n", "")
                    Path(folder_level_3).mkdir(parents=True, exist_ok=True)
                    print(folder_level_3)
                    wrapper_to_folder(cursor,folder_level_3,str(entry[0]))



if __name__ == "__main__":
    main()