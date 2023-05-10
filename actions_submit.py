from ast import keyword
from collections import Counter
from itertools import groupby
from pathlib import Path
from unicodedata import name
from openpyxl import load_workbook
from diff_folder_and_mysql import get_sheet_structure 
import pyodbc 
import os
import shutil



def all_equal(iterable):
    g = groupby(iterable)
    return next(g, True) and not next(g, False)

def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()

def check_exerciseid_in_structure(wb_s,cursor,row,cnxn, structure_file):
    #it checks wrapper and wrappereercise for id and wrapper_id and for exercise_id
    sheet = wb_s['Lessons']
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    min_col, min_row,max_col,max_row=ranges[names.index('WrapperExercises')]
    Create_Entry = False
    Edit_Excel = False
    data = []
    coord = []
    ids = columns[names.index('WrapperExercises')]
    for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
        for cell in cells:
            data.append(cell.value)
            coord.append((cell.row,cell.column))
    print("excel data: ",data, ids, coord)
    
    min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
    for cells in sheet.iter_cols(min_col=min_col,min_row=2, max_col=max_col, max_row=2):
        for cell in cells:
            if cell.value == 'Id':
                Wrapper_Id_column = cell.column
                
    min_col, min_row,max_col,max_row=ranges[names.index('Level names')]
    level = []
    for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
        for cell in cells:
            level.append(cell.value)
    print(level)
    
    level_name = next(item for item in level if item is not None)
    print(level.index(level_name))
    
    finished = False
    level_row = row
    ladder = []
    ladder_val = 1
    ladder_bottom = []
    while not level_row == 3:
        level_row = level_row - 1
        if sheet.cell(row=level_row,column=min_col + level.index(level_name)-ladder_val).value:
            parent = sheet.cell(row=level_row,column=min_col + level.index(level_name)-ladder_val).value
            #if sheet.cell(row=level_row,column=Wrapper_Id_column).value:
            ladder.append((parent,level_row))
            #else:
            #    ladder.append((parent,level_row))
            ladder_val = ladder_val + 1
                
            #    pass
    print(ladder)
    
    #check ladder
    #check entry point
    min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
    entry = []
    for cells in sheet.iter_cols(min_col=min_col,min_row=ladder[-1][-1], max_col=max_col, max_row=ladder[-1][-1]):
        for cell in cells:
            entry.append(cell.value)
    print(entry)
    sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Wrappers]'
    where_match = True
    print(columns[names.index('Wrappers')])
    
    
    for id in columns[names.index('Wrappers')]:
        print(id)
        cell_value = entry[columns[names.index('Wrappers')].index(id)]
        print(cell_value)
        if cell_value == True:
            cell_value = 1
        elif cell_value == False:
            cell_value = 0
        else:
            pass
        if cell_value != None:
            if isinstance(cell_value, str):
                string = '\''+str(cell_value)+'\''
            else:
                string = str(cell_value)
            if where_match == True:
                sqlcommand = sqlcommand + ' WHERE ['+ id + '] = ' + string
                where_match = False
            else: 
                sqlcommand = sqlcommand + ' AND ['+ id + '] = ' + string
        else:
            pass
        
    print(sqlcommand)
    
    cursor.execute(sqlcommand)
    list = cursor.fetchall()
    print(list)
    
    if list:
        pass
    else:
        if entry[columns[names.index('Wrappers')].index('Id')] == None or entry[columns[names.index('Wrappers')].index('Name')] == None or  entry[columns[names.index('Wrappers')].index('Level')] or entry[columns[names.index('Wrappers')].index('RelatedLanguage_Id')] == None:
            print("no entry point info, create")
        else:
            pass

        if entry[columns[names.index('Wrappers')].index('Name')] == None or  entry[columns[names.index('Wrappers')].index('Level')] or entry[columns[names.index('Wrappers')].index('RelatedLanguage_Id')] == None:
            print("not enough info to create the entry point. quitting")
            exit()
        else:
            sqlcommand = 'SELECT MAX(Id) AS maximum FROM Wrappers'

            cursor.execute(sqlcommand)
            grandpa = cursor.fetchall()[0][0]+1
            entry[columns[names.index('Wrappers')].index('Id')] = grandpa
            sqlcommand_insert = 'INSERT INTO [dbo].[Wrappers] VALUES('
            count = 0
            for id in columns[names.index('Wrappers')]:
                if count == 0:
                    sqlcommand_insert = sqlcommand_insert+'?'
                else:
                    sqlcommand_insert = sqlcommand_insert+',?'
                count = count + 1
            sqlcommand_insert = sqlcommand_insert+')'
            print(sqlcommand_insert,entry)
        
            cursor.execute(sqlcommand_insert,entry)
            cnxn.commit()

            min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
            print(min_col, min_row,max_col,max_row)
            count = 0
            for col in range(min_col,max_col+1):
                sheet.cell(row = ladder[-1][-1], column = col).value = entry[count]
                print(count, sheet.cell(row = ladder[-1][-1], column = col).value)
                count = count + 1
            wb_s.save(structure_file)



    



        

    
    grandpa = int(sheet.cell(row=ladder[-1][-1],column=Wrapper_Id_column).value)
    print(grandpa)
    
    for lad in ladder[:-1][::-1]:
        print(lad)
        
        min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
        entry = []
        for cells in sheet.iter_cols(min_col=min_col,min_row=lad[-1], max_col=max_col, max_row=lad[-1]):
            for cell in cells:
                entry.append(cell.value)
        if entry[columns[names.index('Wrappers')].index('WrapperId')] != grandpa:
            print("oops")
            entry[columns[names.index('Wrappers')].index('WrapperId')] = grandpa
            Create_Entry = True
            #sheet.cell(row=lad[-1], column = min_col + columns[names.index('Wrappers')].index('WrapperId')+1).value)
        else:
            pass
        if entry[columns[names.index('Wrappers')].index('Id')] == None:
            Create_Entry = True
        else:
            sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[Wrappers]'
            count = 0
            for id in columns[names.index('Wrappers')]:
                print(id)
                cell_value = entry[columns[names.index('Wrappers')].index(id)]
                print(cell_value)
                if cell_value == True:
                    cell_value = 1
                elif cell_value == False:
                    cell_value = 0
                else:
                    pass
                if cell_value != None:
                    if isinstance(cell_value, str):
                        string = '\''+str(cell_value)+'\''
                    else:
                        string = str(cell_value)
                    if count == 0:
                        sqlcommand = sqlcommand + ' WHERE ['+ id + '] = ' + string
                    else: 
                        if id != 'Name':
                            sqlcommand = sqlcommand + ' AND ['+ id + '] = ' + string
                        else:
                            pass
                else:
                    pass
                count = count + 1
            print(sqlcommand)
            
            cursor.execute(sqlcommand)
            list = cursor.fetchall()
            if list:
                pass
            else:
                Create_Entry= True
            
            
        

        print(Create_Entry)
        
        if Create_Entry == True:
            sqlcommand = 'SELECT MAX(Id) AS maximum FROM Wrappers'

            cursor.execute(sqlcommand)
            grandpa = cursor.fetchall()[0][0]+1
            entry[columns[names.index('Wrappers')].index('Id')] = grandpa
            sqlcommand_insert = 'INSERT INTO [dbo].[Wrappers] VALUES('
            count = 0
            for id in columns[names.index('Wrappers')]:
                if count == 0:
                    sqlcommand_insert = sqlcommand_insert+'?'
                else:
                    sqlcommand_insert = sqlcommand_insert+',?'
                count = count + 1
            sqlcommand_insert = sqlcommand_insert+')'
            print(sqlcommand_insert)
            Edit_Excel = True
            cursor.execute(sqlcommand_insert,entry)
            cnxn.commit()
        else:
            grandpa = entry[columns[names.index('Wrappers')].index('Id')]
        if Edit_Excel:
            min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
            print(min_col, min_row,max_col,max_row)
            count = 0
            for col in range(min_col,max_col+1):
                sheet.cell(row = lad[-1], column = col).value = entry[count]
                print(count, sheet.cell(row = lad[-1], column = col).value)
                count = count + 1
            wb_s.save(structure_file)
            
        else:
            print('No edits to Excel')
        
    #compare wrapper id of an exercise with id of its parent folder from Wrappers columns    
    
    Create_Entry = False

    min_col, min_row,max_col,max_row=ranges[names.index('Wrappers')]
    #search for parent wrapper id scrolling upward
    finished = False
    wrapper_id_row = row
    while not finished:
        wrapper_id_row = wrapper_id_row - 1
        print(sheet.cell(row=wrapper_id_row,column=Wrapper_Id_column).value)
        if sheet.cell(row=wrapper_id_row,column=Wrapper_Id_column).value:
            finished = True
            Wrapper_Id = sheet.cell(row=wrapper_id_row,column=Wrapper_Id_column).value
            if not Wrapper_Id:
                pass
    print(Wrapper_Id, data[ids.index('Wrapper_Id')])
    
    #check where whrapper_id coincides with wrapper id, if not, replace with parents id
    if data[ids.index('Wrapper_Id')] == Wrapper_Id:
        pass
    else:
        sheet.cell(row=coord[ids.index('Wrapper_Id')][0], column=coord[ids.index('Wrapper_Id')][1]).value=Wrapper_Id
    #check if it has an exercise id 
    Create_Entry = False
    if data[ids.index('Exercise_Id')]:
        print('already exists the entry in the excel file, checking for existance in the database')
        sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[WrapperExercises] where Exercise_Id = ' + str(data[ids.index('Exercise_Id')]) + ' AND Wrapper_Id = '+str(Wrapper_Id)
        cursor.execute(sqlcommand)
        list = cursor.fetchall()
        if not list:
            Create_Entry = True
        else:
            print('db entry exists')
            Exercise_Id = data[ids.index('Exercise_Id')]
    else:
        Create_Entry = True

    if Create_Entry == True:
        print('creating an entry')
        cursor.execute('SELECT MAX(Id) AS maximum FROM Exercises')
        Exercise_Id = cursor.fetchall()[0][0]+1
        sqlcommand = 'INSERT INTO [dbo].[WrapperExercises] ([Wrapper_Id],[Exercise_Id]) VALUES '
        list = sqlcommand.split()[3].split(',')
        values = sqlcommand.split()[3].replace('[Wrapper_Id]',str(Wrapper_Id)).replace('[Exercise_Id]',str(Exercise_Id))
        sqlcommand = sqlcommand + values
        cursor.execute(sqlcommand)
        sheet.cell(row=coord[ids.index('Exercise_Id')][0], column=coord[ids.index('Exercise_Id')][1]).value = Exercise_Id
        
    else:
        pass
    return Create_Entry,[Wrapper_Id,Exercise_Id]

def indices(lst, item):
    return [i for i, x in enumerate(lst) if x == item]

def get_entry_by_name(sheet, table_name, names, columns,ranges, row):
    entries = []
    table_indices = indices(names, table_name)
    for index in table_indices:
        entry = []
        styles = []
        min_col, min_row,max_col,max_row=ranges[index]
        range = (min_col, row, max_col, row) 
        for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
            for cell in cells:
                entry.append(cell.value)
                styles.append(type(cell.value))
        entries.append((entry,range,columns[index],styles))
    return entries

def replace_entry(sheet,entry, entry_range):
    min_col, min_row,max_col,max_row=entry_range
    index = 0
    for col in range(min_col,max_col+1):
        cell = sheet.cell(row=min_row, column=col)
        cell.value = entry[index]
        index = index + 1

def work_on_entry(wb, input_col, input_to_local_col, modify_col, Create_Entry, cursor, cnxn,exercise_file, sheet, table_name, row, table_name_number):
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet,table_name=table_name, names=names,ranges=ranges,row=row,columns=columns)[table_name_number]
    
    
    isnone = False
    if all_equal(entry):
        if entry[0] == None:
            isnone = True
        else:
            pass
    else:
        pass
 
    try:
        if entry[entry_columns.index('Key')] == 'ReverseExercise':
            isnone = True
        else:   
            pass
    except:
        pass

    if isnone == False:
        
        Edit_Excel = False
        if input_col == False or input_to_local_col == False:
            pass
        else:
            if entry[entry_columns.index(input_to_local_col)]!= input_col:
                entry[entry_columns.index(input_to_local_col)] = input_col
                Edit_Excel = True
            else:
                pass

        
        if entry[entry_columns.index(modify_col)]:
            sqlcommand = 'SELECT * FROM [CalstContent].[dbo].['+table_name + ']'
            count = 0
            for id in entry_columns:
                cell_value = entry[entry_columns.index(id)]
                if cell_value == True:
                    cell_value = 1
                elif cell_value == False:
                    cell_value = 0
                else:
                    pass
                if cell_value != None:
                    if isinstance(cell_value, str):
                        string = '\''+str(cell_value).replace("'","''")+'\''
                    else:
                        string = str(cell_value)
                    if count == 0:
                        sqlcommand = sqlcommand + ' WHERE ['+ id + '] = ' + string
                    else: 
                        sqlcommand = sqlcommand + ' AND ['+ id + '] = ' + string
                else:
                    pass
                count = count + 1
            print(sqlcommand)
            
            cursor.execute(sqlcommand)
            list = cursor.fetchall()
            print(list)
            
            if not list:
                Create_Entry = True 
                Edit_Excel = True
            else:
                pass
        else:
            Create_Entry = True
            Edit_Excel = True

        if Create_Entry:
            Edit_Excel = True
            sqlcommand = 'SELECT MAX('+modify_col+') AS maximum FROM '+ table_name
            print(sqlcommand)
            cursor.execute(sqlcommand)
            print(entry, "inside create entry")
            modify = cursor.fetchall()[0][0]+1
            if table_name == 'ConfusionBoxes':
                sqlcommand = 'SELECT MAX(ConfusionBox_Id) AS maximum FROM [CalstContent].[dbo].[TranscriptionConfusionBoxes]'
                cursor.execute(sqlcommand)
                max_cb_id = cursor.fetchall()[0][0]+1
                print(modify, max_cb_id)
                if max_cb_id > modify:
                    modify = max_cb_id
                
            print(modify, "modify")
            entry[entry_columns.index(modify_col)] = modify
            sqlcommand_insert = 'INSERT INTO [dbo].['+table_name+'] VALUES('
            count = 0
            for id in entry_columns:
                if count == 0:
                    sqlcommand_insert = sqlcommand_insert+'?'
                else:
                    sqlcommand_insert = sqlcommand_insert+',?'
                count = count + 1
            sqlcommand_insert = sqlcommand_insert+')'
            cursor.execute(sqlcommand_insert,entry)
            print(sqlcommand_insert, entry)
            
            cnxn.commit()
        else:
            print('db entry exists')
        
        if Edit_Excel:
            replace_entry(sheet=sheet,entry=entry,entry_range=entry_range)
            wb.save(str(exercise_file))    
        else:
            print('No edits to Excel')
    else:
        pass
    
        
    return entry, entry_columns

def work_on_entry_with_no_id(wb, input_cols, input_to_local_cols, Create_Entry, cursor, cnxn,exercise_file, sheet, table_name, row):
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet,table_name=table_name, names=names,ranges=ranges,row=row,columns=columns)[0]
    print(entry,'here we are')
    isnone = False
    if all_equal(entry):
        if entry[0] == None:
            isnone = True
        else:
            pass
    else:
        pass
    
    Edit_Excel = False

    for index,input in enumerate(input_to_local_cols):
        if entry[entry_columns.index(input)] != input_cols[index]:
            entry[entry_columns.index(input)] = input_cols[index]
            Edit_Excel = True
        else:
            pass

    
    sqlcommand = 'SELECT * FROM [CalstContent].[dbo].['+table_name + ']'
    count = 0
    for id in entry_columns:
        cell_value = entry[entry_columns.index(id)]
        if isinstance(cell_value, str):
            string = '\''+str(cell_value)+'\''
        else:
            string = str(cell_value)
        if count == 0:
            sqlcommand = sqlcommand + ' WHERE '+ id + ' = ' + string
        else: 
            sqlcommand = sqlcommand + ' AND '+ id + ' = ' + string
        count = count + 1
    cursor.execute(sqlcommand)
    list = cursor.fetchall()
    if not list:
        Create_Entry = True 
        Edit_Excel = True
    else:
        pass
    
    if Create_Entry:
        
        sqlcommand_insert = 'INSERT INTO [dbo].['+table_name+'] VALUES('
        count = 0
        for id in entry_columns:
            if count == 0:
                sqlcommand_insert = sqlcommand_insert+'?'
            else:
                sqlcommand_insert = sqlcommand_insert+',?'
            count = count + 1
        sqlcommand_insert = sqlcommand_insert+')'
        cursor.execute(sqlcommand_insert,entry)
        cnxn.commit()
        
    else:
        print('db entry exists')
    
    if Edit_Excel:
        replace_entry(sheet=sheet,entry=entry,entry_range=entry_range)
        wb.save(str(exercise_file))    
    else:
        print('No edits to Excel')
        
  
    return entry, entry_columns

def work_with_line_in_structure_lessons(line, wb_structure, structure_file, course_sheet, course_path, Force_Rewrite, cursor, cnxn, dst_path):
    path=course_path
    sheet = course_sheet
    folder_col = get_column(sheet=sheet, row=1, name='Folders')
    action_col = get_column(sheet=sheet, row=1, name='Actions')
    cell_action = sheet.cell(row=line,column=action_col)
    
    #structure file changes
    #check wrapper_id exercise_id info, and replace it if needed
    Create_Entry, [Wrapper_Id, Exercise_Id] = check_exerciseid_in_structure(structure_file = structure_file, cnxn=cnxn, wb_s=wb_structure, cursor=cursor, row=line)
    print([Wrapper_Id, Exercise_Id] )
    
    entry_wrex= [Wrapper_Id, Exercise_Id]
    if Create_Entry:
        wb_structure.save(filename=(str(structure_file))) 
        cnxn.commit()
    else:
        pass
   
    
    #open an exercise.xlsx
    exercise_path = Path(sheet.cell(row=line, column=folder_col).value.replace('..',str(path.parent)))
    exercise_file = exercise_path/'exercise.xlsx'
    sounds_folder = exercise_path/'sound_files'
    print(exercise_file.exists(), str(exercise_file))
    
    if sounds_folder.exists():
        for filename in os.listdir(str(sounds_folder)):
            print(filename)
            sound_file =  sounds_folder/filename
            dest_file = dst_path/filename
            print(str(dest_file))
            if os.path.isfile(str(dest_file)):
                print("it exists")
            else:
                shutil.copy(str(sound_file), str(dst_path))
            #if dest_file.exists:
            #    print("it exists")
            #else:
            #shutil.copy(str(sound_file), str(dst_path))
    
    print(str(exercise_file))
   
    wb_exercise = load_workbook(str(exercise_file))

    data_row = 3
    sheet_ex = wb_exercise['Exercise']
    names_ex,columns_ex,ranges_ex = get_sheet_structure(sheet = sheet_ex)
    
    entry_ex,entry_ex_range,entry_ex_columns, entry_ex_styles = get_entry_by_name(sheet=sheet_ex,table_name='WrapperExercises', names=names_ex,ranges=ranges_ex,row=data_row,columns=columns_ex)[0]
    
    #wrapper and exercise ids were defined already in sturcure_lessons.xlsx, so just copy it here if needed
    if entry_ex != entry_wrex:                          
        replace_entry(sheet=sheet_ex,entry=entry_wrex,entry_range=entry_ex_range)
        wb_exercise.save(str(exercise_file))

    #create if needed an exercise entry in the Exercises database
   
    entry_ex, entry_ex_columns = work_on_entry_with_no_id(table_name='Exercises', wb=wb_exercise,
                                        input_cols=[Exercise_Id], 
                                        input_to_local_cols=['Id'], 
                                        Create_Entry=False,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet_ex,row=data_row)
    #there can be more properties entries than 1, so iterate if needed
 
    c_prop = Counter(names_ex)
    print(c_prop)
    
    for i in range(c_prop['Properties']):
            #it creates an entry with proper exercise id and id if it doesnt exist in db, excel file gets upodated, pay attention, 
            # exercises.xlsx shoul be closed
            entry_prop,entry_prop_columns  = work_on_entry(wb=wb_exercise,table_name='Properties',
                                        input_col=Exercise_Id ,input_to_local_col='ExerciseId',modify_col='Id',
                                        Create_Entry=False,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet_ex,row=data_row, table_name_number=i)
        
    
    #next sheet, confusionbox, but we may have vocab confusion box, MP, Nonword, they have different structure, 
    # vocab has one confusionbox id for all words in an exercise, MP - 1 per 2 words, etc
    
    #remove confusionboxes associated with exerciseif if eny
    sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[ConfusionBoxes] where ExerciseId = ' + str(Exercise_Id)
    print(sqlcommand)
    cursor.execute(sqlcommand)
    list = cursor.fetchall()
    print(list)
    if list:
        sqlcommand = 'DELETE FROM [CalstContent].[dbo].[ConfusionBoxes] where ExerciseId = ' + str(Exercise_Id)
        cursor.execute(sqlcommand)
        cnxn.commit()
    else:
        pass

    case_vocab = False
    vocab = []
    case_mp = False
    mp = []
    case_nw = False
    nw = []
    print(wb_exercise.sheetnames, "mitya")
    
    
    for sheetname in wb_exercise.sheetnames:
        print(sheetname)
        if 'Vocab' in  sheetname:
            case_vocab = True
            vocab.append(sheetname)
        elif 'MP' in sheetname:
            #work on Vocab-Confusion Box
            case_mp = True
            mp.append(sheetname)
        elif 'NonWords' in sheetname:
            case_nw = True
            nw.append(sheetname)
        else:
            pass
  
    print(case_mp,case_vocab, case_nw)
    
    
    if case_vocab:
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = False
        print(vocab)
        max_info = []
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in vocab:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in vocab if 'Confusion' in s][0]
        print(sheet_name)
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in vocab if 'Words Properties' in s][0]
        print(sheet_name)
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
                
        
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        
        for row in range(data_row,maximum_row+1):
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab    
            if first_run == True:
                first_run = False
                #it checks and possibly creates an entry in the db, excel file is getting modified as well, if db is updated. 
                # make sure an excel file is closed. 
                entry_cb,entry_cb_columns  = work_on_entry(wb=wb_exercise,table_name='ConfusionBoxes',
                                        input_col=Exercise_Id ,input_to_local_col='ExerciseId',modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet,row=row, table_name_number=0)                       
            else:
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet,table_name='ConfusionBoxes', names=names,ranges=ranges,row=row,columns=columns)[0]
                if entry != entry_cb:                          
                    replace_entry(sheet=sheet,entry=entry_cb,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))

            #words, either creates a new entry, or leave as it is, depending on Force_Rewrite value            
            entry_word, entry_word_columns = work_on_entry(table_name='Words', wb=wb_exercise,
                                        input_col=False, input_to_local_col=False, modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet,row=row,table_name_number=0) 
            
            
            
            #transcriptions
            #for Spanish, Greak and Italian there are 2 speakers and thus two transcriptions fileds, but for Norwegian there could be mulriple, 
            # which looks wrong a bit
            c = Counter(names)
            print(c['Transcriptions'])
            
            for num in range(c['Transcriptions']):
                #transcription, wordid and id columns
                entry_trans, entry_trans_columns = work_on_entry(table_name='Transcriptions', wb=wb_exercise,
                                            input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row,table_name_number=num)
                print(entry_trans, entry_trans_columns) 
                if num == 0:
                    print(num, "here")
                    #transcriptionconfusionboxes binds two Speakers with a confusionbox id, via just one transcription id
                    print(entry_cb, entry_trans, "here we are")
                    
                    entry_cb_trans, entry_cb_trans_columns = work_on_entry_with_no_id(table_name='TranscriptionConfusionBoxes', wb=wb_exercise,
                                            input_cols=[entry_trans[entry_trans_columns.index('Id')],entry_cb[entry_cb_columns.index('Id')]], 
                                            input_to_local_cols=['Transcription_Id','ConfusionBox_Id'], 
                                            Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row)
                    
                
                    print(entry_cb_trans, entry_cb_trans_columns)
                    
                else:
                    pass
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in vocab if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Words', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_word)
                
                if entry != entry_word:                          
                    replace_entry(sheet=sheet_st,entry=entry_word,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))

                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Transcriptions', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_trans)
                
                
                if entry != entry_trans:                          
                    replace_entry(sheet=sheet_st,entry=entry_trans,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                for pron_num in range(c_pron['Pronunciations']):
                    print(pron_num)
                    #transcription id and id should be correct, they are checked, and be either replaced or not,
                    #  depends on existence and Force_rewrite
                    entry_pron, entry_pron_columns = work_on_entry(table_name='Pronunciations', wb=wb_exercise,
                                            input_col=entry_trans[entry_trans_columns.index('Id')], input_to_local_col='Transcription_Id', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet_st,row=row,table_name_number=pron_num) 
            
            #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[0]
            if entry != entry_word:  
                print(entry, entry_word)                        
                replace_entry(sheet=sheet_wp,entry=entry_word,entry_range=entry_range)
                wb_exercise.save(str(exercise_file))
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                print(i,"here")
                entry_prop, entry_prop_columns = work_on_entry(table_name='Properties', wb=wb_exercise,
                                        input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet_wp,row=row,table_name_number=i)
                print(entry_prop, entry_prop_columns)
            
    if case_mp:
        
        print(mp)
        
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        #Force_Rewrite = True
        
        max_info = []
        sheets = mp
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in sheets:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
      
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in sheets if 'Confusion' in s][0]
      
        
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)

        sheet_name = [s for s in sheets if 'Words Properties' in s][0]
     
        
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
                
        
        max_word = row
       
        if maximum_row != max_word:
            maximum_row = max_word
        
       
        
        
        for row in range(data_row,maximum_row+1):
            
            #start with confusionbox,
            # this will either create a new confusionbox id or just replace current entry with the same confusionbox id details, 
            # just one entry for all the lines in case of vocab 
               
            if first_run == True:
                
                
                first_run = False
                #it checks and possibly creates an entry in the db, excel file is getting modified as well, if db is updated. 
                # make sure an excel file is closed. 
                entry_cb,entry_cb_columns  = work_on_entry(wb=wb_exercise,table_name='ConfusionBoxes',
                                        input_col=Exercise_Id ,input_to_local_col='ExerciseId',modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet,row=row, table_name_number=0) 
                print(entry_cb)
                                   
            else:
                
                entry,entry_range,entry_columns,entry_styles = get_entry_by_name(sheet=sheet,table_name='ConfusionBoxes', names=names,ranges=ranges,row=row,columns=columns)[0]
               
                if entry != entry_cb:                          
                    replace_entry(sheet=sheet,entry=entry_cb,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))
                first_run = True   
                
                


            #words, either creates a new entry, or leave as it is, depending on Force_Rewrite value            
            entry_word, entry_word_columns = work_on_entry(table_name='Words', wb=wb_exercise,
                                        input_col=False, input_to_local_col=False, modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet,row=row,table_name_number=0) 
            
            
            
            #transcriptions
            #for Spanish, Greak and Italian there are 2 speakers and thus two transcriptions fileds, but for Norwegian there could be mulriple, 
            # which looks wrong a bit
            c = Counter(names)
            print(c['Transcriptions'])
            
            for num in range(c['Transcriptions']):
                #transcription, wordid and id columns
                entry_trans, entry_trans_columns = work_on_entry(table_name='Transcriptions', wb=wb_exercise,
                                            input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row,table_name_number=num)
                #print(entry_trans, entry_trans_columns) 
                if num == 0:
                    print(num, "here")
                    #transcriptionconfusionboxes binds two Speakers with a confusionbox id, via just one transcription id
                    print(entry_cb, entry_trans, "here we are")
                    
                    entry_cb_trans, entry_cb_trans_columns = work_on_entry_with_no_id(table_name='TranscriptionConfusionBoxes', wb=wb_exercise,
                                            input_cols=[entry_trans[entry_trans_columns.index('Id')],entry_cb[entry_cb_columns.index('Id')]], 
                                            input_to_local_cols=['Transcription_Id','ConfusionBox_Id'], 
                                            Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row)
                    
                
                    print(entry_cb_trans, entry_cb_trans_columns)
                    
                else:
                    pass
                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in sheets if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Words', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_word)
                
                if entry != entry_word:                          
                    replace_entry(sheet=sheet_st,entry=entry_word,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))

                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Transcriptions', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_trans)
                
                
                if entry != entry_trans:                          
                    replace_entry(sheet=sheet_st,entry=entry_trans,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                for pron_num in range(c_pron['Pronunciations']):
                    print(pron_num)
                    #transcription id and id should be correct, they are checked, and be either replaced or not,
                    #  depends on existence and Force_rewrite
                    entry_pron, entry_pron_columns = work_on_entry(table_name='Pronunciations', wb=wb_exercise,
                                            input_col=entry_trans[entry_trans_columns.index('Id')], input_to_local_col='Transcription_Id', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet_st,row=row,table_name_number=pron_num) 
            
            #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[0]
            if entry != entry_word:  
                print(entry, entry_word)                        
                replace_entry(sheet=sheet_wp,entry=entry_word,entry_range=entry_range)
                wb_exercise.save(str(exercise_file))
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                print(i,"here")
                entry_prop, entry_prop_columns = work_on_entry(table_name='Properties', wb=wb_exercise,
                                        input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet_wp,row=row,table_name_number=i)
                print(entry_prop, entry_prop_columns, "here last last")
                

    if case_nw:
        print(nw, "here")
        
        #confusionbox id should be created or checked once, all other lines carry the same confusionbox id, that's whe the keyword firstrun 
        # was introduced
        first_run = True
        Force_Rewrite = True
        
        max_info = []
        sheets = nw
        #this checks whether all sheets have the same number of lines as it should be. Consistency check.
        for sheet_name in sheets:
            sheet = wb_exercise[sheet_name]
            max_info.append(sheet.max_row)
        print(max_info)
        
        isequal = all_equal(max_info)
        if isequal:
            maximum_row = max_info[0]
        else:
            print('something wrong with number of rows, terminating')
            exit()
        
        #info lines are 2 rows now, but I keep the possibility that we may have more, that's why I use data_row which is always three as a parameter, 
        # for future changes
        data_row = 3
        #sheet names for word properties and confusion box, to be used next steps
        sheet_name = [s for s in sheets if 'Confusion' in s][0]
        print(sheet_name)
        
        sheet = wb_exercise[sheet_name]
        names,columns,ranges = get_sheet_structure(sheet = sheet)
        print(names)
        
        sheet_name = [s for s in sheets if 'Words Properties' in s][0]
        print(sheet_name)
        
        sheet_wp = wb_exercise[sheet_name]
        names_wp,columns_wp,ranges_wp = get_sheet_structure(sheet = sheet_wp)
        
        min_col, min_row, max_col, max_row = ranges_wp[names_wp.index('Words')]
        for row in range(3,sheet_wp.max_row+1):
            cell = sheet_wp.cell(row=row,column=min_col+columns_wp[names_wp.index('Words')].index('Text'))
            if cell.value == None:
                break
                
        
        max_word = row
        print(max_word, maximum_row)
        if maximum_row != max_word:
            maximum_row = max_word
        
        print(maximum_row)
        
        
        for row in range(data_row,maximum_row+1):
            
            #start with confusionbox,
               
            
            #it checks and possibly creates an entry in the db, excel file is getting modified as well, if db is updated. 
            # make sure an excel file is closed. 
            entry_cb,entry_cb_columns  = work_on_entry(wb=wb_exercise,table_name='ConfusionBoxes',
                                    input_col=Exercise_Id ,input_to_local_col='ExerciseId',modify_col='Id',
                                    Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                    sheet=sheet,row=row, table_name_number=0) 
            print(entry_cb)
                             
            
            #words, either creates a new entry, or leave as it is, depending on Force_Rewrite value            
            entry_word, entry_word_columns = work_on_entry(table_name='Words', wb=wb_exercise,
                                        input_col=False, input_to_local_col=False, modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet,row=row,table_name_number=0) 
            
            #properties
            print(entry_cb[0], entry_cb_columns)
            print(names)
            
            if 'Properties' in names:
                entry_nwprop, entry_nwprop_columns = work_on_entry(table_name='Properties', wb=wb_exercise,
                                            input_col=str(entry_cb[entry_cb_columns.index('Id')]), input_to_local_col='Description', modify_col='Id',
                                            Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row,table_name_number=0) 
            else:
                pass            
            
            #transcriptions
            #for Spanish, Greak and Italian there are 2 speakers and thus two transcriptions fileds, but for Norwegian there could be mulriple, 
            # which looks wrong a bit
            c = Counter(names)
            print(c['Transcriptions'])
            
            for num in range(c['Transcriptions']):
                #transcription, wordid and id columns
                entry_trans, entry_trans_columns = work_on_entry(table_name='Transcriptions', wb=wb_exercise,
                                            input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row,table_name_number=num)
                print(entry_trans, entry_trans_columns) 
                if num == 0:
                    print(num, "here")
                    #transcriptionconfusionboxes binds two Speakers with a confusionbox id, via just one transcription id
                    print(entry_cb, entry_trans, "here we are")
                    
                    entry_cb_trans, entry_cb_trans_columns = work_on_entry_with_no_id(table_name='TranscriptionConfusionBoxes', wb=wb_exercise,
                                            input_cols=[entry_trans[entry_trans_columns.index('Id')],entry_cb[entry_cb_columns.index('Id')]], 
                                            input_to_local_cols=['Transcription_Id','ConfusionBox_Id'], 
                                            Create_Entry=True,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet,row=row)
                    
                
                    print(entry_cb_trans, entry_cb_trans_columns)
                    
                else:
                    pass




                #here we go to speaker transciptrion sheets, where sound files for speakers are specified
                sheet_name = [s for s in sheets if 'Speaker_Trans_'+str(num) in s][0]
                sheet_st = wb_exercise[sheet_name]
                names_st,columns_st,ranges_st = get_sheet_structure(sheet = sheet_st)
                #words and transcription are borrowed from previous confusionbox sheet
                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Words', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_word)
                
                if entry != entry_word:                          
                    replace_entry(sheet=sheet_st,entry=entry_word,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))

                entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_st,table_name='Transcriptions', names=names_st,ranges=ranges_st,row=row,columns=columns_st)[0]
                print(entry, entry_trans)
                
                
                if entry != entry_trans:                          
                    replace_entry(sheet=sheet_st,entry=entry_trans,entry_range=entry_range)
                    wb_exercise.save(str(exercise_file))
                
                #can be several sound files per word, so here we iterate over all occurances
                c_pron = Counter(names_st)
                for pron_num in range(c_pron['Pronunciations']):
                    print(pron_num)
                    #transcription id and id should be correct, they are checked, and be either replaced or not,
                    #  depends on existence and Force_rewrite
                    entry_pron, entry_pron_columns = work_on_entry(table_name='Pronunciations', wb=wb_exercise,
                                            input_col=entry_trans[entry_trans_columns.index('Id')], input_to_local_col='Transcription_Id', 
                                            modify_col='Id',Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                            sheet=sheet_st,row=row,table_name_number=pron_num) 
            
            #the next sheet, Words, where we specify properties of words, normally it is translations and pictures, if any

            #words entry is borrowed from a confusionbox sheet
            entry,entry_range,entry_columns, entry_styles = get_entry_by_name(sheet=sheet_wp,table_name='Words', names=names_wp,ranges=ranges_wp,row=row,columns=columns_wp)[0]
            if entry != entry_word:  
                print(entry, entry_word)                        
                replace_entry(sheet=sheet_wp,entry=entry_word,entry_range=entry_range)
                wb_exercise.save(str(exercise_file))
            
            c = Counter(names_wp)
            print(c['Properties'], "here")
            #can be more than 1 properties, iterate
            
            for i in range(c['Properties']):
                print(i,"here")
                entry_prop, entry_prop_columns = work_on_entry(table_name='Properties', wb=wb_exercise,
                                        input_col=entry_word[entry_word_columns.index('Id')], input_to_local_col='WordId', modify_col='Id',
                                        Create_Entry=Force_Rewrite,cursor=cursor,cnxn=cnxn,exercise_file=exercise_file,
                                        sheet=sheet_wp,row=row,table_name_number=i)
                print(entry_prop, entry_prop_columns)
            

    else:
        pass

               
def main(course_folder,cursor, cnxn, keyword, output_sounds):
    path = Path(course_folder)
    #the path to the course summary file
    structure_file  = Path(course_folder+'\\lessons_structure.xlsx')

    if structure_file.exists():
        #check actions column for the command "submit"
        wb_structure = load_workbook(str(structure_file))
        sheet = wb_structure['Lessons']
        to_submit = []
        print(str(structure_file))
        action_col = get_column(sheet=sheet, row = 1, name='Actions')
        
        for cells in sheet.iter_cols(min_col=action_col,min_row=3, max_col=action_col, max_row=sheet.max_row):
            for cell in cells:
                print(str(cell.value).lower())
                if str(cell.value).lower() == keyword: #or str(cell.value).lower() == 'submitted' or str(cell.value).lower() == 'retract' or str(cell.value).lower() == 'submitted-nonword' or str(cell.value).lower() == 'failed':
                    to_submit.append(cell.row)
        # list to_submit contains rows of summary file to submit
        print('rows to_submit: ', to_submit)
        
        
        for line in to_submit:
            cell_action = sheet.cell(row=line,column=action_col)
            #try:
            work_with_line_in_structure_lessons(line=line,wb_structure=wb_structure,cursor=cursor,cnxn=cnxn,
                                                    structure_file=structure_file,course_sheet=sheet,course_path=path, Force_Rewrite=False, dst_path=output_sounds)
            cell_action.value = 'submitted'
            wb_structure.save(structure_file)
            #except:
                #cell_action.value = 'failed_submitted'
                #wb_structure.save(structure_file)  
                     
        
    else:
        print("No exercise file here. quitting")
        exit() 

if __name__ == "__main__":
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    
    #connect to the calst database
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    language='Spanish'
    output_sound = r'C:\Source\Repos\CalstEnglish\CalstFiles\WordObjectContent'+'\\'+language+r'\OriginalWords_Wav'
    print(output_sound)
 
    keyword = 'retract'	
    dst_path = Path(output_sound)
    #folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\'
    #folder = 'G:\\My Drive\\CALST_courses\\'+str(language)+'_course_styled\\'
    #folder = 'C:\\Source\\Repos\\mysql-excel\\'+str(language)+'_course_styled\\'
    folder = r'C:\Users\dmitrysh\OneDrive - NTNU\CALST_courses\\'+str(language)+'_course_styled'
    #folder = r'C:\Users\dmitrysh\OneDrive - NTNU'
    main(course_folder=folder, cursor=cursor, cnxn=cnxn, keyword=keyword, output_sounds = dst_path)
    cnxn.commit()
    cnxn.close()