import pyodbc 
import argparse
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils.cell import cols_from_range, coordinate_to_tuple,get_column_interval
from diff_folder_and_mysql import get_sheet_structure

def missing_in_excel(excel):
    firstrun = True
    first_pron = True
    wb = load_workbook(filename = excel)
    meta_accum = []
    header_cells = []
    for i in wb.sheetnames:
        sheet = wb[i]
        
        if 'speaker' in i.lower():
            header = ''
            print(i)
            fails = []
            failed_row = []
            names,columns,ranges = get_sheet_structure(sheet = sheet)
            print(columns)
            
            for row in range(2,sheet.max_row):
                first_pron = True
                meta = []
                for data in zip(names,columns,ranges):
                    (col_low, row_low, col_high, row_high) = data[-1]
                    entry = []
                    for cells in sheet.iter_cols(min_col=col_low,min_row=row_low+row, max_col=col_high, max_row= row_low+row):
                        for cell in cells:
                            entry.append(cell.value)
                    meta.append(entry)
                    if data[0] == 'Pronunciations' and first_pron == True:
                        first_pron = False
                        if all(x is None for x in entry):
                            fails.append(row)
                            failed_row.append(meta)
                            break
                        else:
                            pass
                    else:
                        pass

            if fails: 
                accum = []
                for fail in zip(fails,failed_row):
                    line = [i]
                    line.append(fail[0])
            
                    for item in fail[1]:
                        line  = line + item
                    accum.append(line)
                
                if firstrun == True:
                    firstrun = False
                    header = ['sheet','row']
                    header_cells = [1,1]
                    for index,col in enumerate(columns):
                        if index < 3:
                            print(col)
                            header = header + col
                            header_cells.append(len(col))
                    print(header_cells)
                    meta_accum.append(header)
                 
                else:
                    pass
                meta_accum = meta_accum+accum
            else:
                pass

        else:
            pass
      
    return meta_accum,header_cells     

 
def main(folder,output_diff):
    folder_path = Path(folder)
    
    folder_path = Path(folder)
    print(str(folder_path.parent))
    
    if folder_path.exists():
        pass
    else:
        print(' There is no such folder as "'+folder+'"','\n Please enter the correct folder name')
        exit()
  
    excel = folder_path/'exercise.xlsx'
    output = missing_in_excel(excel=excel)
    for i in output[0]:
        print(i)
  

if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='python diff_folder_and_mysql.py -f foldername')
    parser.add_argument('-f',dest='folder')
    parser.add_argument('-diff',dest='diff',default='missing.txt')
    args = parser.parse_args()
    if args.folder:
        main(folder = args.folder, diff=args.diff)
    else:
        #folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\The alphabet'
        folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\Beginner\\Lesson 10\\uo\\uo-o_backslash_u'
        main(folder=folder, output_diff='difft.txt')