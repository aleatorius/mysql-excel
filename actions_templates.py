from natsort import natsorted
from itertools import groupby
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import diff_folder_and_mysql
import course_structure_db_excel
from make_template import template_out_of_exercise


def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()

def get_style(header_cells,sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            if cell.row > 1:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))+2.7
                except:
                    pass
            else:
                pass
        adjusted_width = (max_length) 
        sheet.column_dimensions[column].width = adjusted_width   
    start=1
    twocolor = ['00CCFFCC','00FFFF99','00C0C0C0','0033CCCC']
    ccolor = twocolor[0]
    twopattern = ['lightDown','darkGray']
    for i in header_cells:
        
        cell_header = sheet.cell(row=1, column=start)
        double = Side(border_style="double", color="00008000")
        cell_header.border = Border(top=double, left=double, right=double, bottom=double)
        cell_header.font  = Font(b=True, color="00008000", size = 8)
        cell_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False,  shrink_to_fit=True)
    
        #for j in range(i):
        #    cell_h = sheet.cell(row=1, column=start+j)
        #    cell_h.font  = Font(b=True, color="00008000", size = 10)
        #    cell_h.alignment = Alignment(horizontal="general", vertical="bottom", wrap_text=False,  shrink_to_fit=True)
        #    if j == 0:
        #        cell_h.border = Border(left=double, bottom=double)
        #    else:
        #        cell_h.border = Border(bottom=double)

        
        ppattern = twopattern[0]
        for l in range(2,sheet.max_row+1):
            for j in range(i):
                
                cell_ordinary = sheet.cell(row = l,column=start+j)
                cell_ordinary.fill = PatternFill(ppattern, fgColor=ccolor)
                thin = Side(border_style="thin", color="000000")
                cell_ordinary.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if ppattern == twopattern[0]:
                ppattern = twopattern[1]
            else:
                ppattern = twopattern[0]
        
        start = start + i
        if ccolor == twocolor[0]:
            ccolor = twocolor[1]
        else:
            ccolor = twocolor[0]
    
    #if header_cells[-1] == 1:
    #    cell_h.border = Border(bottom=double, right=double,left=double)   
    #else:
    #    cell_h.border = Border(bottom=double, right=double)     

def main(folder):
    path = Path(folder)
    print(path.parent)
    structure_path  = Path(folder+'\\lessons_structure.xlsx')
    firstrun = True
    if structure_path.exists():
        wb_s = load_workbook(str(structure_path))
        sheet = wb_s.active
        to_diff = []
        print(str(structure_path))
        
        exercise_col = get_column(sheet=sheet,row=2,name='Exercise_Id')
        folder_col = get_column(sheet=sheet,row=1,name='Folders')
        print(exercise_col,folder_col)
        for cells in sheet.iter_cols(min_col=exercise_col,min_row=3, max_col=exercise_col, max_row=sheet.max_row):
            for cell in cells:
                print(cell.value)
                if cell.value:
                    if cell.value != None:
                        cell_f = sheet.cell(column=folder_col, row=cell.row)
                        print(cell_f.value)
                        path_folder = str(path.parent)+cell_f.value.replace('..','')
                        to_diff.append(path_folder)
        #print(to_diff)
        
        for folder in to_diff[3:]:
            print(folder)
            try:
                template_out_of_exercise(folder=folder)
            except:
                pass
            
            
  
if __name__ == "__main__":

    folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\'
    #folder = 'G:\\My Drive\\CALST_courses\\Spanish_course_styled\\'
    main(folder=folder)