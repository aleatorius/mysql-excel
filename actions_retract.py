from pathlib import Path
from openpyxl import load_workbook
from diff_folder_and_mysql import get_sheet_structure 
import pyodbc 



def remove_exerciseid_in_structure(wb_s,cursor,row,cnxn, structure_file):
    print(row)

    Edit_Excel = False
    #it checks wrapper and wrappereercise for id and wrapper_id and for exercise_id
    sheet = wb_s['Lessons']
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    min_col, min_row,max_col,max_row=ranges[names.index('WrapperExercises')]
    entry= []
    for cells in sheet.iter_cols(min_col=min_col,min_row=row, max_col=max_col, max_row=row):
        for cell in cells:
            entry.append(cell.value)
    if all(x is None for x in entry):
        pass
    else:
        sqlcommand = 'SELECT * FROM [CalstContent].[dbo].[WrapperExercises]'
        count = 0
        for id in columns[names.index('WrapperExercises')]:
            print(id)
            cell_value = entry[columns[names.index('WrapperExercises')].index(id)]
            if cell_value == True:
                cell_value = 1
            elif cell_value == False:
                cell_value = 0
            else:
                pass
            if cell_value != None:
                if isinstance(cell_value, str):
                    string = '\''+str(cell_value)+'\''
                else:
                    string = str(cell_value)
                if count == 0:
                    sqlcommand = sqlcommand + ' WHERE ['+ id + '] = ' + string
                else: 
                    sqlcommand = sqlcommand + ' AND ['+ id + '] = ' + string
            else:
                pass
            count = count + 1


        cursor.execute(sqlcommand)
        print(sqlcommand, "before list")
        list = cursor.fetchall()
        print(list, "list")
        
        if list:
            sqlcommand = 'DELETE FROM [CalstContent].[dbo].[WrapperExercises]'
            count = 0
            for id in columns[names.index('WrapperExercises')]:
                print(id)
                cell_value = entry[columns[names.index('WrapperExercises')].index(id)]
                print(cell_value, "here")
                if cell_value == True:
                    cell_value = 1
                elif cell_value == False:
                    cell_value = 0
                else:
                    pass
                if cell_value != None:
                    if isinstance(cell_value, str):
                        string = '\''+str(cell_value)+'\''
                    else:
                        string = str(cell_value)
                    if count == 0:
                        sqlcommand = sqlcommand + ' WHERE ['+ id + '] = ' + string
                    else: 
                        sqlcommand = sqlcommand + ' AND ['+ id + '] = ' + string
                else:
                    pass
                count = count + 1
            print(sqlcommand)
            exit()
            cursor.execute(sqlcommand)
            cnxn.commit()
            Edit_Excel = True
                
        else:
            pass
            

        
        if Edit_Excel:
            min_col, min_row,max_col,max_row=ranges[names.index('WrapperExercises')]
            print(min_col, min_row,max_col,max_row)
            count = 0
            for col in range(min_col,max_col+1):
                print(count, sheet.cell(row = row, column = col).value)
                sheet.cell(row = row, column = col).value = None
                print(count, sheet.cell(row = row, column = col).value)
                count = count + 1
        
            wb_s.save(structure_file)
            
        else:
            print('No edits to Excel')
        
        #compare wrapper id of an exercise with id of its parent folder from Wrappers columns    



def get_column(sheet, row, name):
    col = []
    for cells in sheet.iter_cols(min_col=1,min_row=row, max_col=sheet.max_column, max_row=row):
        for cell in cells:
            if cell.value == name:
                col.append(cell.column)
    if len(col) == 1:
        return col[0]
    else:
        print('Warning! Several cols with name '+name+ ' exiting')
        exit()

    
def indices(lst, item):
    return [i for i, x in enumerate(lst) if x == item]
           
def main(course_folder,cursor, cnxn):
    path = Path(course_folder)
    #the path to the course summary file
    structure_file  = Path(course_folder+'\\lessons_structure.xlsx')

    if structure_file.exists():
        #check actions column for the command "submit"
        wb_structure = load_workbook(str(structure_file))
        sheet = wb_structure['Lessons']
        to_submit = []
        print(str(structure_file))
        action_col = get_column(sheet=sheet, row = 1, name='Actions')
        
        for cells in sheet.iter_cols(min_col=action_col,min_row=3, max_col=action_col, max_row=sheet.max_row):
            for cell in cells:
                print(str(cell.value).lower())
                if str(cell.value).lower() == 'resubmit' or str(cell.value).lower() == 'submit' or str(cell.value).lower() == 'retract':
                    to_submit.append(cell.row)
        # list to_submit contains rows of summary file to submit
        print('rows to_submit: ', to_submit)
        
        
        for line in to_submit:
            print(line)
            remove_exerciseid_in_structure(wb_s=wb_structure,cursor=cursor,row=line,cnxn=cnxn, structure_file=structure_file)
           
    else:
        print("No exercise file here. quitting")
        exit() 

if __name__ == "__main__":
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    
    #connect to the calst database
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    #folder = 'C:\\Source\\Repos\\mysql-excel\\Spanish_course_styled\\'
    folder = 'G:\\My Drive\\CALST_courses\\Italian_course_styled\\'
    main(course_folder=folder, cursor=cursor, cnxn=cnxn)
    cnxn.commit()
    cnxn.close()