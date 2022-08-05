from site import execsitecustomize
import pyodbc 
import argparse
from itertools import groupby

from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils.cell import cols_from_range, coordinate_to_tuple,get_column_interval

def all_equal(iterable):
    g = groupby(iterable)
    return next(g, True) and not next(g, False)

def compare(entry, data):
    if len(entry) == len(data):
        diff = []
        for ind,ex in enumerate(entry):
            if ex == data[ind]:
                diff.append(True)
            else:
                diff.append(False)
        return diff
    else:
        print("ERROR in compare entries", entry, data)
            
def get_full_request(cursor, database, data, entry):
    sqlcom = 'SELECT * FROM ['+database+'].[dbo].[' + data[0]+'] WHERE '
    count = 0
    for ind, ex in enumerate(data[1]):
        if count < len(entry)-1:
            sqlcom = sqlcom + str(ex) + ' = ' + str(entry[ind]) + ' AND '
        else:
            sqlcom = sqlcom + str(ex) + ' = ' + str(entry[ind]) 
        count = count + 1 
    cursor.execute(sqlcom)
    output = [list(i) for i in cursor.fetchall()] 
    return output    

def get_sheet_structure(sheet):
    #a database row specified as merged cells
    merged_ranges = sheet.merged_cells.ranges
    database = []
    ranges = []
    for range in merged_ranges:
        (col_low, row_low, col_high, row_high) = range.bounds
        ranges.append(range.bounds)
        database.append(sheet.cell(row_low,col_low).value)
    database_and_ranges = zip(database, ranges)
    database_cols = []
    for data in database_and_ranges:
        (col_low, row_low, col_high, row_high) = data[-1]
        if row_high-row_low != 0:
            print("Error,too many rows for info caption")
            exit()
        else:
            pass
        
        col_id = []
        for row in sheet.iter_rows(min_col=col_low,max_col=col_high,min_row=row_low+1,max_row=row_high+1):
            for cell in row:
                col_id.append(cell.value)
        database_cols.append(col_id)

    database_ranges_columns = zip(database,database_cols, ranges)
    return database, database_cols,ranges

def diff_write(diff,diff_file,output,entry,data,sheet,row):
    indices = [i for i, x in enumerate(diff) if x == False]
    (min_col, min_row, max_col, max_row) = data[-1]
    temp = get_column_interval(min_col, max_col)
    for index in indices:
        if not output[index]:
            print(output[index],"is empty")
        else:
            pass
        diff_file.write('Sheet: "'+sheet.title+ '" Cell:"'+ str(temp[index])+str(min_row+row)+' ' + str(data[1][index]) +'"\n Excel: "'+ str(entry[index]) +'" db: "'+ str(output[index])+'"\n')

def noentry_write(noentry_file,entry,data,sheet,row):
    (min_col, min_row, max_col, max_row) = data[-1]
    temp = get_column_interval(min_col, max_col)
    noentry_file.write('Sheet: "'+sheet.title+ '" Cell:"'+ str(temp)+str(min_row+row)+' ' + str(data[1]) +'"\n Excel: "'+ str(entry) +'\n')

def get_columns_to_compare(sheet, id_set_to_compare):
    print(id_set_to_compare)
    if type(id_set_to_compare) is tuple:
        id_set_to_compare = [id_set_to_compare]
    elif type(id_set_to_compare) is list:
        pass
    else:
        print('neither a tuple or a list')
        exit()
    db_row = 1
    db_names,db_columns,db_ranges = get_sheet_structure(sheet = sheet)
    output = []
    for data in zip(db_names,db_columns,db_ranges):
        (col_low, row_low, col_high, row_high) = data[-1]
        for cols in sheet.iter_cols(min_col=col_low,min_row=row_low+db_row, max_col=col_high, max_row= row_low+db_row):
            for cell in cols:
                for col in id_set_to_compare:
                    if cell.value == col[0] and data[0] == col[1]:
                        output.append(cell.column)
    return output


def compare_between_values_in_columns(sheet,input_columns,warnings_file):
    
    db_cols_row = 2
    output = []
    for row in range(db_cols_row+1,sheet.max_row+1):
        columns_to_compare = []
        compared_cells_cols = []
        for col in input_columns:
            cell = sheet.cell(row=row,column=col)
            if cell.value:
                columns_to_compare.append(cell.value)
                compared_cells_cols.append(cell.column_letter)
            else:
                pass
        if all_equal(columns_to_compare) == True:
            output.append(True)
        else:
            output.append(False)
            warnings_file.write("ERROR, column values do not match! Check these cells and correct them:\n")
            string = 'Sheet: '+sheet.title+' '
            for cell_col in compared_cells_cols:
                string = string + cell_col + str(row)+' '
            warnings_file.write(string + '\n')
            warnings_file.write('Values: '+str(columns_to_compare)+'\n\n')
    return output


def check_exerciseid(sheet,input_columns,warnings_file):
    db_cols_row = 2
    row=db_cols_row+1
    columns_to_compare = []
    compared_cells_cols = []
    for col in input_columns:
        cell = sheet.cell(row=row,column=col)
        columns_to_compare.append(cell.value)
        compared_cells_cols.append(cell.column_letter)
    if all_equal(columns_to_compare) == True:
        columns_to_compare = [columns_to_compare[0]]
    else:
        warnings_file.write("ERROR, exercise ids do not match! Check these cells and correct them:\n")
        string = ''
        for cell_col in compared_cells_cols:
            string = string + cell_col + str(row)+' '
        warnings_file.write(string + '\n')
        warnings_file.write('Values: '+str(columns_to_compare)+'\n\n')
    return columns_to_compare 



def check_one_value_columns_vs_value(sheet,input_columns,value, warnings_file):
    db_cols_row = 2
    fails = []
    for colval in input_columns:  
        for row in range(db_cols_row+1,sheet.max_row+1):
            cell = sheet.cell(row=row,column=colval)
            if str(cell.value) != str(value):
                fails.append((cell.column_letter+str(row)+' '+str(cell.value)))
            
    if fails:
        warnings_file.write("ERROR, error in the column! Check this column and correct it:\n"+'Sheet: '+sheet.title+'\n')
        for fail in fails:
            warnings_file.write(fail+' should be '+str(value)+'\n')
        warnings_file.write('\n')
        
        


def check_one_value_columns(sheet,input_columns,warnings_file):
    db_cols_row = 2
    columns_to_compare = [] 
    columns_letter =  []
    for colval in input_columns:              
        for row in range(db_cols_row+1,sheet.max_row+1):
            cell = sheet.cell(row=row,column=colval)
            columns_to_compare.append(cell.value)
        columns_letter.append(cell.column_letter)
    print(columns_to_compare)
    if all_equal(columns_to_compare) == True:
        return True
    else:

        warnings_file.write("ERROR, error in the columns! Check these columns and correct them:\n")
        warnings_file.write('Sheet: ' +sheet.title+' Rows: '+ str(columns_letter)+'\n\n')
        return False




def main(folder):
    
    warnings_file = open('warnings.txt','w')
    noentry_file = open('noentry.txt', 'w')
    ser_file = open('server_private.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    
    #connect to the calst database
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    
    folder_path = Path(folder)
    if folder_path.exists():
        pass
    else:
        print(' There is no such folder as "'+folder+'"','\n Please enter the correct folder name')
        exit()
    
    exercise = folder_path/'exercise.xlsx'
   
    if exercise.exists():
        print("ok, an exercise exists in this folder")
        wb = load_workbook(filename = exercise)
        
        sheet = wb['Exercise']
        
        if sheet.max_row != 3:
            warnings_file.write('For the sheet: '+sheet.title+' there are more rows than 3: '+ str(sheet.max_row)+'\n')
        if sheet.max_column < 14:
            warnings_file.write('For the sheet: '+sheet.title+' there are less than 14 columns: '+ str(sheet.max_column)+'\n')
       
        #check wrapper for this exercise
        
        id_set_to_compare = [('Exercise_Id','WrapperExercises'),('Id','Exercises'),('ExerciseId','Properties')]
        exerciseid_cols = get_columns_to_compare(sheet=sheet, id_set_to_compare=id_set_to_compare)
        compared = check_exerciseid(sheet=sheet,input_columns=exerciseid_cols,warnings_file=warnings_file)
        #it will output an exercise id if it is the same 
        if len(compared) == 1:
            ExerciseId = compared[0]
        else:
            print(compared, "error, serious mismatch, quitting")
            exit()
        
        
        #check exercises in the different sheet, compare with exercised
        
        confusionbox_sheet =  []
        matches = ['ConfusionBox']
        for name in wb.sheetnames:
            if any(x in name for x in matches):
                confusionbox_sheet.append(name)
       
        for cb in confusionbox_sheet:
            sheet_cb = wb[cb]
            id_set_to_compare = [('ExerciseId','ConfusionBoxes')]
            #check exercise_id with respect to the sheet "Exercise"
            exerciseid_cols = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=id_set_to_compare)
            check_one_value_columns_vs_value(sheet=sheet_cb,input_columns=exerciseid_cols,value=ExerciseId, warnings_file=warnings_file)
            
            #check confusionbox ids
            id_set_to_compare = [('Id','ConfusionBoxes'),('ConfusionBox_Id','TranscriptionConfusionBoxes')]
            confusionboxid_cols = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=id_set_to_compare)
            check_one_value_columns(sheet=sheet_cb,input_columns=confusionboxid_cols, warnings_file=warnings_file)
            
            id_set_to_compare = [('Transcription_Id','TranscriptionConfusionBoxes'),('Id','Transcriptions')]
            transcriptions_cols = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=[id_set_to_compare[-1]])
            
            #attention -- bad place -- if more then one transcription we need to choose the first one
            if len(transcriptions_cols) == 2:
                wholeset  = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=id_set_to_compare)[0:2]
                compared = compare_between_values_in_columns(sheet=sheet_cb,input_columns=wholeset,warnings_file=warnings_file)
            elif len(transcriptions_cols) == 1:
                wholeset  = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=id_set_to_compare)
                compared = compare_between_values_in_columns(sheet=sheet_cb,input_columns=wholeset,warnings_file=warnings_file)
            else:
                print("error confusionbox block")
                exit()
            id_set_to_compare = [('Id','Words'),('WordId','Transcriptions')]
            wordid = get_columns_to_compare(sheet=sheet_cb, id_set_to_compare=id_set_to_compare)
            compared = compare_between_values_in_columns(sheet=sheet_cb,input_columns=wordid,warnings_file=warnings_file)
        
        wordproperties_sheet =  []
        matches = ['Words Properties']
        for name in wb.sheetnames:
            if any(x in name for x in matches):
                wordproperties_sheet.append(name)
        for wp in wordproperties_sheet :
            sheet_wp = wb[wp]   
            id_set_to_compare = [('Id','Words'), ('WordId','Properties')]
            wordid_wp = get_columns_to_compare(sheet=sheet_wp, id_set_to_compare=id_set_to_compare)
            print(wordid_wp)
            compared = compare_between_values_in_columns(sheet=sheet_wp,input_columns=wordid_wp,warnings_file=warnings_file)
  
        speaker_sheet =  []
        matches = ['Speaker']

        for name in wb.sheetnames:
            if any(x in name for x in matches):
                speaker_sheet.append(name)
        print(speaker_sheet)
        for sp in speaker_sheet :
            sheet_sp = wb[sp]
            id_set_to_compare = [('Transcription_Id','Pronunciations'), ('Id','Transcriptions')]
            trans_sp = get_columns_to_compare(sheet=sheet_sp, id_set_to_compare=id_set_to_compare)
            compared = compare_between_values_in_columns(sheet=sheet_sp,input_columns=trans_sp,warnings_file=warnings_file)
            
    else:
        print('No excel file in this folder')
    warnings_file.close
    ser_file.close
    noentry_file.close      

if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='python check_db_vs_excel_consistency.py -f foldername')
    parser.add_argument('-f',dest='folder')
    args = parser.parse_args()
    if args.folder:
        main(folder = args.folder)
    else:
        #folder='C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\The alphabet'
        folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\\Numbers 1'
        main(folder=folder)