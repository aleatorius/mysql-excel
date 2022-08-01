from tkinter import E
import pyodbc 
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

def get_sheet_structure(sheet):
    #a database row specified as merged cells
    merged_ranges = sheet.merged_cells.ranges
    database = []
    ranges = []
    for range in merged_ranges:
        (col_low, row_low, col_high, row_high) = range.bounds
        ranges.append(range.bounds)
        database.append(sheet.cell(row_low,col_low).value)
    database_and_ranges = zip(database, ranges)
    database_cols = []
    for data in database_and_ranges:
        (col_low, row_low, col_high, row_high) = data[-1]
        col_id = []
        for row in sheet.iter_rows(min_col=col_low,max_col=col_high,min_row=row_low+1,max_row=row_high+1):
            for cell in row:
                col_id.append(cell.value)
        database_cols.append(col_id)

    database_ranges_columns = zip(database,database_cols, ranges)
    return database_ranges_columns
        
def main():
    folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 3\The house'
    folder_path = Path(folder)
    exercise = folder_path/'exercise.xlsx'
    if exercise.exists():
        print("ok, exercise exists in this folder")
        wb = load_workbook(filename = exercise)
        for i in wb.sheetnames:
            print(i)
            structure = get_sheet_structure(sheet = wb[i])
            for element in structure:
                print(i,": ",element)
            
                
            


           
if __name__ == "__main__":
    main()