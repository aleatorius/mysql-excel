from ctypes import Structure
from textwrap import indent
from natsort import natsorted
import pyodbc 
import os
import argparse
from itertools import groupby
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from diff_folder_and_mysql import get_sheet_structure

def get_wrapper(wrapper):
    wb = load_workbook(filename = wrapper)
    print(wrapper)
    sheet = wb['Wrappers']
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    print(names)

    data_row = sheet.max_row
    entry = []
    for data in zip(names,columns,ranges):
        (col_low, row_low, col_high, row_high) = data[-1]
        if data[0] == 'Wrappers':
            for cells in sheet.iter_cols(min_col=col_low,min_row=data_row, max_col=col_high, max_row=data_row):
                for cell in cells:
                    entry.append(cell.value)
        else:
            pass
    index = columns[0].index('Name')
    level = (entry,entry[index],names,columns)
    return level

def get_exercise(wrapper):
    level = []
    wb = load_workbook(filename = wrapper)
    sheet = wb['Exercise']
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    data_row = sheet.max_row
    id_col_row = 2
    exercise_names = []
    entry = []
    cols=[]
    for data in zip(names,columns,ranges):
        if data[0] == 'WrapperExercises':
            (col_low, row_low, col_high, row_high) = data[-1]
            for cells in sheet.iter_cols(min_col=col_low,min_row=data_row, max_col=col_high, max_row=data_row):
                for cell in cells:
                    entry.append(cell.value)
            header_ex_id = data[1]
            header_ex = len(header_ex_id)*[data[0]]
            
        if data[0] == 'Properties':
            (col_low, row_low, col_high, row_high) = data[-1]
            
            
            for cells in sheet.iter_cols(min_col=col_low,min_row=id_col_row, max_col=col_high, max_row=id_col_row):
                for cell in cells:
                    if cell.value == 'Key':
                        key = cell.col_idx
                    if cell.value == 'Value':
                        value =  cell.col_idx 

            data_cell = sheet.cell(column=key,row=data_row)
            if data_cell.value == 'ExerciseName':
                name_cell = sheet.cell(column=value,row=data_row)
                exercise_names.append(name_cell.value)
            
        

    if len(exercise_names) == 0 or len(exercise_names)>1:
        print("error, too many names or none, quitting")
        exit()
    else:
        return entry,exercise_names[0],header_ex_id, header_ex
    
    return level

def get_wrapper_dirs(folder):
    local_depth = os.path.abspath(folder).count(os.path.sep)
    local_level = []
    for subdir, dirs, files in os.walk(folder, topdown=True):
        for name in files:
            filepath = subdir + os.sep + name
            if filepath.endswith(".xlsx"):
                if name == 'wrapper.xlsx':
                    if subdir.count(os.path.sep) == local_depth:
                        local_level = dirs
    if local_level:
        return local_level
    else:
        return False

def get_style(header_cells,sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            if cell.row > 1:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))+2.7
                except:
                    pass
            else:
                pass
        adjusted_width = (max_length) 
        sheet.column_dimensions[column].width = adjusted_width   
    start=1
    twocolor = ['00CCFFCC','00FFFF99','00C0C0C0','0033CCCC']
    ccolor = twocolor[0]
    twopattern = ['lightDown','darkGray']
    for i in header_cells:
        cell_header = sheet.cell(row=1, column=start)
        double = Side(border_style="double", color="00008000")
        cell_header.border = Border(top=double, left=double, right=double, bottom=double)
        cell_header.font  = Font(b=True, color="00008000", size = 8)
        cell_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False,  shrink_to_fit=False)
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start+i-1)
        for j in range(i):
            cell_h = sheet.cell(row=2, column=start+j)
            cell_h.font  = Font(b=True, color="00008000", size = 10)
            cell_h.alignment = Alignment(horizontal="general", vertical="bottom", wrap_text=False,  shrink_to_fit=False)
            if j == 0:
                cell_h.border = Border(left=double, bottom=double)
            else:
                cell_h.border = Border(bottom=double)

        
        ppattern = twopattern[0]
        for l in range(3,sheet.max_row+1):
            for j in range(i):
                cell_ordinary = sheet.cell(row = l,column=start+j)
                cell_ordinary.fill = PatternFill(ppattern, fgColor=ccolor)
                thin = Side(border_style="thin", color="000000")
                cell_ordinary.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if ppattern == twopattern[0]:
                ppattern = twopattern[1]
            else:
                ppattern = twopattern[0]
        
        start = start + i
        if ccolor == twocolor[0]:
            ccolor = twocolor[1]
        else:
            ccolor = twocolor[0]
         
    if header_cells[-1] == 1:
        cell_h.border = Border(bottom=double, right=double,left=double)   
    else:
        cell_h.border = Border(bottom=double, right=double)   
    
    names,columns,ranges = get_sheet_structure(sheet = sheet)
    (col_low, row_low, col_high, row_high) = ranges[names.index('Folders')]
       
    for cells in sheet.iter_cols(min_col=col_low,min_row=row_low+2, max_col=col_high, max_row=sheet.max_row):
        for cell in cells:
            if cell.value:
                cell.hyperlink = cell.value
    
                        
    



def main(folder, output):
    inputfolder = folder
    warnings_file = open('warnings.txt','w')
    ser_file = open('server_private_remote.md','r')
    info = []
    for i in ser_file:
        info.append(i.split()[-1].replace("'",''))
    [server,database,username,password] = info
    
    #connect to the calst database
    #cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+'; UID='+username+';PWD='+ password)
    #cursor = cnxn.cursor()
    
    folder_path = Path(folder)
    if folder_path.exists():
        pass
    else:
        print(' There is no such folder as "'+folder+'"','\n Please enter the correct folder name')
        exit()
    abspath = os.path.abspath(folder_path)
    
    wbout = Workbook()
    sheet_wbout = wbout.active
    sheet_wbout.title = "Lessons"
    header_basic = ['Level 0','Level 1', 'Level 2', 'Level 3','Level 4','Add/Del/Mod/Diff']
    
    header_cells = []
    header = header_basic
    header_0 = (len(header)-1)*['Level names']
    header_0.append('Actions')
    header_cells.append(len(header)-1)
    header_cells.append(1)
    
    name_structure=[]
    wrapper_structure = []
    structure = []
    folders = []
    
    #entry level
    wrapper = folder_path/'wrapper.xlsx'
    level=get_wrapper(wrapper=wrapper)
    header_0_wrapper = len(level[3][0])*level[2]    
    header_wrapper = level[3][0]
    header_cells.append(len(level[3][0]))
    newline = []
    newline.append(level[1])
    name_structure.append(newline)
    wrapper_structure.append(level[0])
    root = 'C:\Source\Repos\mysql-excel'
    folders.append(os.path.abspath(folder_path).replace(root,'..'))
    
 
    first_time = True
    level = get_wrapper_dirs(folder_path)
    for chapter in level:
        print(chapter)
        newline = ['']
        folder = folder_path/chapter
        lessons = get_wrapper_dirs(folder)
        wrapper = folder/'wrapper.xlsx'
        level=get_wrapper(wrapper=wrapper)
        newline.append(level[1])
        folders.append(os.path.abspath(folder).replace(root,'..'))
        name_structure.append(newline)
        wrapper_structure.append(level[0])
        
        for lesson in natsorted(lessons):
            folder = folder_path/chapter/lesson
            wrapper = folder/'wrapper.xlsx'
            level=get_wrapper(wrapper=wrapper)
            newline = 2*['']
            newline.append(level[1])
            folders.append(os.path.abspath(folder).replace(root,'..'))
            name_structure.append(newline)
            wrapper_structure.append(level[0])
            sublessons = get_wrapper_dirs(folder)
            for sublesson in natsorted(sublessons):
                folder=folder_path/chapter/lesson/sublesson
                sub_sublessons = get_wrapper_dirs(folder)
                if sub_sublessons != False:
                    folder=folder_path/chapter/lesson/sublesson
                    wrapper = folder/'wrapper.xlsx'
                    level=get_wrapper(wrapper=wrapper)

                    newline = 3*['']
                    newline.append(level[1])
                    folders.append(os.path.abspath(folder).replace(root,'..'))
                    name_structure.append(newline)
                    wrapper_structure.append(level[0])

                    for sub_sublesson in sub_sublessons:
                        structure.append((chapter, lesson,sublesson,sub_sublesson))
                        folder = folder_path/chapter/lesson/sublesson/sub_sublesson
                        wrapper = folder/'exercise.xlsx'
                        level = get_exercise(wrapper=wrapper)
                        if first_time == True:
                            header_ex_id = level[2]
                            header_ex = level[3] 
                            header_cells.append(len(header_ex))
                            first_time = False
                        else:
                            pass
                        newline = 4*['']
                        newline.append(level[1])
                        folders.append(os.path.abspath(folder).replace(root,'..'))
                        name_structure.append(newline)
                        wrapper_structure.append(level[0])
                else:
                    structure.append((chapter,lesson,sublesson))
                    folder = folder_path/chapter/lesson/sublesson
                    wrapper = folder/'exercise.xlsx'
                    if wrapper.exists():
                        level = get_exercise(wrapper=wrapper)
                        if first_time == True:
                            header_ex_id = level[2]
                            header_ex = level[3] 
                            header_cells.append(len(header_ex))
                            first_time = False
                        else:
                            pass
                        newline = 3*['']
                        newline.append(level[1])
                        folders.append(os.path.abspath(folder).replace(root,'..'))
                        wrapper_structure.append(level[0])
                        name_structure.append(newline)
                    else:
                        pass
    header_0 = header_0 + header_0_wrapper+header_ex+['Folders']
    header = header + header_wrapper+header_ex_id+['Relative path']
    header_cells = header_cells + [1]
    sheet_wbout.append(header_0)
    sheet_wbout.append(header)                
    for line,info,fl in zip(name_structure,wrapper_structure,folders):
        for ind,ex in enumerate(line):
            if ex != 0:
                nonzero_index = ind
        if len(info) > 2:
            linen = line+(len(header_basic)-ind-1)*['']+info+len(header_ex)*['']+[fl]
        else:
            linen = line+(len(header_basic)-ind-1+len(header_wrapper))*['']+info+[fl]
        sheet_wbout.append(linen)
    get_style(header_cells=header_cells,sheet=sheet_wbout)
    wbout.save(inputfolder+'\\'+output)
    wbout.close()        
    

  
if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='python course_structure_db_excel.py -f foldername')
    parser.add_argument('-f',dest='folder')
    parser.add_argument('-o',dest='output')
    args = parser.parse_args()
    if args.folder:
        main(folder = args.folder, output=args.output)
    else:
        #folder='C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\The alphabet'
        #folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 1\\Numbers 1'
        #folder = 'C:\Source\Repos\python_tools\Spanish_course_styled\Beginner\Lesson 2\\Nationalities'
        #folder = 'C:\Source\Repos\mysql-excel\Greek_course_styled'
        folder = r'C:\Source\Repos\mysql-excel\English_course_revised'
        output = 'lessons_structure.xlsx'
        main(folder=folder,output= output)